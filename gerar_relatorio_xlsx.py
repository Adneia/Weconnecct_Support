import json, re
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

# ── Helpers ────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def normaliza_motivo(m):
    if not m:
        return 'Não informado'
    m2 = m.strip().lower()
    if re.search(r'atraso|atrasp|attraso|atarso', m2):
        return 'Atraso na entrega'
    if re.search(r'desconhece|n.o reconhece', m2):
        return 'Desconhece a entrega'
    if re.search(r'avaria', m2):
        return 'Avaria'
    if re.search(r'endere.o n.o loc|endere.o ins', m2):
        return 'Endereço não localizado/insuf.'
    if re.search(r'extravio|roubo|perda', m2):
        return 'Extravio / Roubo'
    if re.search(r'item inc|item err|item div|item falt|item a mais', m2):
        return 'Item incorreto / faltante'
    if re.search(r'devolu', m2):
        return 'Devolução'
    if re.search(r'defeito|vício', m2):
        return 'Defeito / Vício'
    if re.search(r'troca de etiq', m2):
        return 'Troca de etiqueta'
    return m.strip()

# ── Cores por categoria ────────────────────────────────────────────────────
CORES = {
    'Falha Compras':   {'h': 'FFF0E0', 'hdr': 'C0392B', 'tab': 'E74C3C'},
    'Falha Produção':  {'h': 'E8F4FD', 'hdr': '1A5276', 'tab': '2980B9'},
    'Falha Transporte':{'h': 'EAF5EA', 'hdr': '1E8449', 'tab': '27AE60'},
}

# ── Carrega dados ──────────────────────────────────────────────────────────
with open('C:/WeConnect/chamados_relatorio.json', encoding='utf-8') as f:
    docs = json.load(f)

hoje_str = datetime.now().strftime('%d/%m/%Y')
hoje_dt  = datetime.now()

# Normaliza categoria e motivo
for d in docs:
    cat = d.get('categoria', '')
    if cat in ('Falha Compras', 'Falha de Compras'):
        d['cat'] = 'Falha Compras'
    elif cat == 'Falha Produção':
        d['cat'] = 'Falha Produção'
    else:
        d['cat'] = 'Falha Transporte'
    d['motivo_norm'] = normaliza_motivo(d.get('motivo', ''))

    try:
        da = d.get('data_abertura','')
        if da:
            dt = datetime.fromisoformat(da.replace('Z','+00:00')).replace(tzinfo=None)
            d['dias'] = (hoje_dt - dt).days
        else:
            d['dias'] = ''
    except:
        d['dias'] = ''

cats_ordem = ['Falha Compras', 'Falha Produção', 'Falha Transporte']
por_cat    = {c: [d for d in docs if d['cat'] == c] for c in cats_ordem}

# ── Workbook ───────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove sheet padrão

# ══════════════════════════════════════════════════════════════════════════
# ABA RESUMO
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('Resumo')
ws.sheet_properties.tabColor = '2C3E50'

title_font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
sub_font    = Font(name='Arial', bold=True, size=10, color='2C3E50')
label_font  = Font(name='Arial', size=9)
value_font  = Font(name='Arial', size=9, bold=True)
gray_fill   = PatternFill('solid', start_color='2C3E50', end_color='2C3E50')
light_gray  = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')
center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
right_al    = Alignment(horizontal='right',  vertical='center')

# Título principal
ws.merge_cells('A1:H1')
ws['A1'] = f'RELATÓRIO DE FALHAS — WECONNECT   |   {hoje_str}'
ws['A1'].font      = title_font
ws['A1'].fill      = gray_fill
ws['A1'].alignment = center
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:H2')
ws['A2'] = 'Falha Compras · Falha Produção · Falha Transporte'
ws['A2'].font = Font(name='Arial', italic=True, size=9, color='7F8C8D')
ws['A2'].alignment = center
ws.row_dimensions[2].height = 16

# Totais por categoria
ws.row_dimensions[4].height = 20
ws['A4'] = 'Categoria'
ws['B4'] = 'Total'
ws['C4'] = 'Pendentes'
ws['D4'] = 'Sem pendência'
ws['E4'] = '< 15 dias'
ws['F4'] = '15–30 dias'
ws['G4'] = '30–60 dias'
ws['H4'] = '> 60 dias'

for col in range(1, 9):
    c = ws.cell(row=4, column=col)
    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='2C3E50', end_color='2C3E50')
    c.alignment = center
    c.border = thin_border()

row = 5
for cat in cats_ordem:
    lst = por_cat[cat]
    total     = len(lst)
    pend      = sum(1 for d in lst if d.get('pendente'))
    sem_pend  = total - pend
    fx15      = sum(1 for d in lst if isinstance(d['dias'], int) and d['dias'] < 15)
    f1530     = sum(1 for d in lst if isinstance(d['dias'], int) and 15 <= d['dias'] < 30)
    f3060     = sum(1 for d in lst if isinstance(d['dias'], int) and 30 <= d['dias'] < 60)
    fgt60     = sum(1 for d in lst if isinstance(d['dias'], int) and d['dias'] >= 60)
    cor_fill  = PatternFill('solid', start_color=CORES[cat]['h'], end_color=CORES[cat]['h'])
    for col, val in enumerate([cat, total, pend, sem_pend, fx15, f1530, f3060, fgt60], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = Font(name='Arial', size=9, bold=(col==1))
        c.fill   = cor_fill
        c.alignment = center if col > 1 else left_al
        c.border = thin_border()
    row += 1

# Linha totais
total_all = len(docs)
fill_tot  = PatternFill('solid', start_color='D5D8DC', end_color='D5D8DC')
for col, val in enumerate(['TOTAL', f'=B5+B6+B7', f'=C5+C6+C7',
                            f'=D5+D6+D7', f'=E5+E6+E7',
                            f'=F5+F6+F7', f'=G5+G6+G7', f'=H5+H6+H7'], 1):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = Font(name='Arial', bold=True, size=9)
    c.fill   = fill_tot
    c.alignment = center if col > 1 else left_al
    c.border = thin_border()

row += 2

# Top 8 motivos por categoria (lado a lado)
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = 'TOP MOTIVOS POR CATEGORIA'
ws[f'A{row}'].font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
ws[f'A{row}'].fill = gray_fill
ws[f'A{row}'].alignment = center
ws.row_dimensions[row].height = 18
row += 1

col_starts = [1, 3, 6]
for idx, cat in enumerate(cats_ordem):
    cstart = col_starts[idx]
    ws.merge_cells(start_row=row, start_column=cstart, end_row=row,
                   end_column=cstart + (1 if idx < 2 else 2))
    hdr_cell = ws.cell(row=row, column=cstart, value=cat)
    hdr_cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    hdr_cell.fill = PatternFill('solid', start_color=CORES[cat]['hdr'],
                                end_color=CORES[cat]['hdr'])
    hdr_cell.alignment = center

row += 1
for idx, cat in enumerate(cats_ordem):
    cstart = col_starts[idx]
    for lbl, txt in [(cstart, 'Motivo'), (cstart+1, 'Qtd')]:
        c = ws.cell(row=row, column=lbl, value=txt)
        c.font = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='566573', end_color='566573')
        c.alignment = center

row_motivo_start = row + 1
counters = {cat: Counter(d['motivo_norm'] for d in por_cat[cat]) for cat in cats_ordem}
top8     = {cat: counters[cat].most_common(8) for cat in cats_ordem}
max_rows = max(len(v) for v in top8.values())

for r_off in range(max_rows):
    fill_r = PatternFill('solid', start_color='F8F9FA' if r_off % 2 == 0 else 'FFFFFF',
                         end_color='F8F9FA' if r_off % 2 == 0 else 'FFFFFF')
    for idx, cat in enumerate(cats_ordem):
        cstart = col_starts[idx]
        if r_off < len(top8[cat]):
            motivo, cnt = top8[cat][r_off]
        else:
            motivo, cnt = '', ''
        for col, val in [(cstart, motivo), (cstart+1, cnt)]:
            c = ws.cell(row=row_motivo_start + r_off, column=col, value=val)
            c.font      = Font(name='Arial', size=8)
            c.fill      = fill_r
            c.alignment = center if col > cstart else left_al
            c.border    = thin_border()

# Larguras resumo
for col, w in [(1, 22), (2, 8), (3, 10), (4, 12), (5, 10), (6, 10), (7, 10), (8, 10)]:
    set_col_width(ws, col, w)

# ══════════════════════════════════════════════════════════════════════════
# ABAS POR CATEGORIA
# ══════════════════════════════════════════════════════════════════════════
COLUNAS = [
    ('ATD',           'id_atendimento', 14),
    ('Data Abertura', 'data_abertura',  14),
    ('Dias',          'dias',            7),
    ('Parceiro/Canal','parceiro',       16),
    ('Nº Pedido',     'numero_pedido',  14),
    ('Cliente',       'nome_cliente',   22),
    ('NF',            'nota_fiscal',    10),
    ('Filial',        'filial',          8),
    ('Motivo',        'motivo_norm',    22),
    ('Status Pedido', 'status_pedido',  28),
    ('Pendente',      'pendente',       10),
    ('Mot. Pendência','motivo_pendencia',20),
    ('Atendente',     'atendente',      16),
    ('Anotações',     'anotacoes',      40),
]

for cat in cats_ordem:
    tab_name = cat.replace('Falha ', '')   # Compras / Produção / Transporte
    ws2 = wb.create_sheet(tab_name)
    ws2.sheet_properties.tabColor = CORES[cat]['tab']
    ws2.freeze_panes = 'A3'

    # Título
    ws2.merge_cells(f'A1:{get_column_letter(len(COLUNAS))}1')
    ws2['A1'] = f'{cat.upper()}   —   {hoje_str}   ({len(por_cat[cat])} atendimentos)'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws2['A1'].fill      = PatternFill('solid', start_color=CORES[cat]['hdr'],
                                      end_color=CORES[cat]['hdr'])
    ws2['A1'].alignment = center
    ws2.row_dimensions[1].height = 22

    # Cabeçalho
    for col_idx, (label, _, width) in enumerate(COLUNAS, 1):
        c = ws2.cell(row=2, column=col_idx, value=label)
        c.font      = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        c.fill      = PatternFill('solid', start_color='2C3E50', end_color='2C3E50')
        c.alignment = center
        c.border    = thin_border()
        set_col_width(ws2, col_idx, width)
    ws2.row_dimensions[2].height = 18

    # Dados
    lst = sorted(por_cat[cat],
                 key=lambda d: d.get('dias', 0) if isinstance(d.get('dias'), int) else 0,
                 reverse=True)

    for r_idx, doc in enumerate(lst, 3):
        fill_row = PatternFill('solid',
                               start_color='F5F5F5' if r_idx % 2 == 0 else 'FFFFFF',
                               end_color='F5F5F5' if r_idx % 2 == 0 else 'FFFFFF')
        for col_idx, (_, field, _) in enumerate(COLUNAS, 1):
            val = doc.get(field, '')
            if field == 'data_abertura' and val:
                try:
                    val = datetime.fromisoformat(val.replace('Z','+00:00')).strftime('%d/%m/%Y')
                except:
                    pass
            if field == 'pendente':
                val = 'Sim' if val else 'Não'
            if val is None:
                val = ''
            c = ws2.cell(row=r_idx, column=col_idx, value=val)
            c.font      = Font(name='Arial', size=8)
            c.fill      = fill_row
            c.border    = thin_border()
            c.alignment = center if col_idx in (3, 7, 8, 11) else left_al

        # Destaca dias > 30 em laranja suave
        if isinstance(doc.get('dias'), int) and doc['dias'] > 30:
            ws2.cell(row=r_idx, column=3).font = Font(name='Arial', size=8,
                                                      bold=True, color='C0392B')
        ws2.row_dimensions[r_idx].height = 15

OUTPUT = 'C:/Users/DELL/Downloads/Relatorio_Falhas_26-05-2026.xlsx'
wb.save(OUTPUT)
print(f'Salvo em: {OUTPUT}')
print(f'Total de atendimentos: {len(docs)}')
for cat in cats_ordem:
    print(f'  {cat}: {len(por_cat[cat])}')
