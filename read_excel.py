import openpyxl, json

wb = openpyxl.load_workbook('/tmp/cancelamento_geral.xlsx', read_only=True, data_only=True)

print('=== SHEETS ===')
for name in wb.sheetnames:
    print(f'  "{name}"')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== SHEET: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===')
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5:
            break
        rows.append(list(row))
    for r in rows:
        print(r)

wb.close()
