"""
Lê o Excel do Emergent, pega todos com reversa,
atualiza o MongoDB do ELO Claude E o Google Sheet.
"""
import asyncio, sys, os
sys.path.insert(0, '/app')
os.chdir('/app')
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from google.oauth2.service_account import Credentials
import gspread

SPREADSHEET_ID = '1cqzY_i1lqvu8sySPFrMtucQfyTo1LYm04ZpxRZNDCBs'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
EXCEL_PATH = '/tmp/atendimentos.xlsx'

async def sync():
    db = AsyncIOMotorClient('mongodb://elo-mongo:27017')['elo_weconnect']

    # --- Lê o Excel ---
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Entrega=0, Reversa=9 (zero-indexed)
    idx_entrega = headers.index('Entrega')
    idx_reversa = headers.index('Reversa')

    # Monta dict entrega -> reversa (só não-vazios)
    excel_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        entrega = str(row[idx_entrega] or '').strip().replace('.0', '')
        reversa = str(row[idx_reversa] or '').strip()
        if entrega and reversa:
            excel_data[entrega] = reversa

    print(f'Excel: {len(excel_data)} entradas com reversa')

    # --- Atualiza MongoDB ---
    all_docs = await db.chamados.find(
        {}, {'id_atendimento': 1, 'numero_pedido': 1, 'codigo_reversa': 1, '_id': 0}
    ).to_list(10000)

    # Índice por numero_pedido
    by_pedido = {str(d.get('numero_pedido', '')).strip(): d for d in all_docs}

    mongo_updates = {}  # entrega -> nova reversa (para confirmar o que mudou)
    for entrega, reversa in excel_data.items():
        doc = by_pedido.get(entrega)
        if not doc:
            continue
        existing = str(doc.get('codigo_reversa') or '').strip()
        if existing == reversa:
            continue
        aid = doc.get('id_atendimento')
        await db.chamados.update_one(
            {'id_atendimento': aid},
            {'$set': {'codigo_reversa': reversa}}
        )
        mongo_updates[entrega] = {'aid': aid, 'reversa': reversa}

    print(f'MongoDB: {len(mongo_updates)} chamados atualizados com reversa')

    # --- Atualiza Google Sheet ---
    creds = Credentials.from_service_account_file('/app/credentials.json', scopes=SCOPES)
    client = gspread.authorize(creds)
    ws_sheet = client.open_by_key(SPREADSHEET_ID).sheet1
    all_values = ws_sheet.get_all_values()
    headers_sheet = all_values[0]

    # Encontra coluna Reversa e coluna Entrega na planilha
    try:
        col_reversa = headers_sheet.index('Reversa') + 1  # 1-indexed
        col_entrega = headers_sheet.index('Entrega') + 1
    except ValueError:
        print('ERRO: Coluna Reversa ou Entrega não encontrada na planilha')
        return

    # Monta índice entrega -> row_number
    entrega_to_row = {}
    for i, row in enumerate(all_values[1:], start=2):
        entrega_val = str(row[col_entrega - 1] or '').strip().replace('.0', '')
        if entrega_val:
            entrega_to_row[entrega_val] = i

    # Prepara batch updates para o Sheet
    sheet_updates = []
    for entrega, reversa in excel_data.items():
        row_num = entrega_to_row.get(entrega)
        if not row_num:
            continue
        # Verifica valor atual na planilha
        current_val = str(all_values[row_num - 1][col_reversa - 1] or '').strip()
        if current_val == reversa:
            continue
        sheet_updates.append({
            'range': f'{chr(64 + col_reversa)}{row_num}',
            'values': [[reversa]]
        })

    if sheet_updates:
        ws_sheet.batch_update(sheet_updates)
        print(f'Google Sheet: {len(sheet_updates)} células atualizadas')
    else:
        print('Google Sheet: nenhuma célula precisou de atualização')

    print(f'\n=== CONCLUÍDO ===')
    print(f'MongoDB: {len(mongo_updates)} atualizados')
    print(f'Google Sheet: {len(sheet_updates)} atualizados')

asyncio.run(sync())
