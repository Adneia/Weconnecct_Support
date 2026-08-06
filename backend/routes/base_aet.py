"""
Módulo: Base AET (Análise de Estoque para Tratativa)
Importa o xlsx que vem do time de Compras com decisões de cancelamento/similar.

Aba esperada: "Analise"
Colunas relevantes:
- [0]  Entrega
- [3]  Produto
- [4]  SKU
- [5]  Cód. Fornecedor
- [6]  Fornecedor
- [7]  Quantidade Solicitada
- [8]  Valor Total do Item
- [11] Pedido de Compra
- [12] Canal de Vendas
- [13] Filial
- [14] Status (do tabelão)
- [15] Data Status
- [16] Dias Parado
- [17] Ação
- [18] Retorno
- [19] Decisão (Cancelar | Oferecer similar | ...)
- [20] Status Atual
- [21] Concluído (Sim/Não)

Regra de decisão:
- Decisão = "Cancelar"                          → acao = "Cancelar"
- Decisão = "Oferecer similar" E Retorno NÃO tem "sem similar" → acao = "Similar"
- Decisão = "Oferecer similar" E Retorno tem "sem similar"     → acao = "Cancelar"
  (fornecedor não encontrou substituto, vira cancelamento)
"""
import io
import uuid
import logging
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, Depends, UploadFile, File
from utils.auth import get_current_user
from utils.database import db

router = APIRouter()
logger = logging.getLogger(__name__)


def _norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def _to_date_iso(v):
    """YYYY-MM-DD"""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _classifica_decisao(decisao_raw: str) -> str:
    """
    Retorna 'Cancelar', 'Similar' ou '' (desconsiderar).
    Segue diretamente a coluna Decisão (T) da AET:
    - 'Cancelar'                         → Cancelar
    - 'Oferecer similar' / 'Substituir'  → Similar
    - branco / outros                    → ignorar
    """
    d = _norm(decisao_raw)
    if not d:
        return ""
    if d.startswith("cancelar"):
        return "Cancelar"
    if "similar" in d or "substituir" in d:
        return "Similar"
    return ""  # ignora finalizado/vazio/etc


@router.post("/api/base-aet/importar")
async def importar_base_aet(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Importa o xlsx da AET. Lê a aba 'Analise' e cria cancelamentos AES
    apenas para pedidos NOVOS (que ainda não estão no banco).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"ok": False, "message": "openpyxl não instalado"}

    # Trava anti-concorrência: duas importações sobrepostas duplicavam os AES
    # (a 2ª carregava os "existentes" antes de a 1ª gravar — caso real de 13/07,
    # 71 pares). Trava expira sozinha em 15 min (caso um import morra no meio).
    _agora = datetime.now(timezone.utc)
    _lock = await db.import_locks.find_one({"chave": "aet"}, {"_id": 0})
    if _lock and _lock.get("em_andamento"):
        try:
            _ini = datetime.fromisoformat(_lock.get("iniciado_em", ""))
            if (_agora - _ini).total_seconds() < 900:
                return {"ok": False, "message": "Já existe uma importação AET em andamento — aguarde ela terminar (ou tente de novo em alguns minutos)."}
        except Exception:
            pass
    await db.import_locks.update_one(
        {"chave": "aet"},
        {"$set": {"em_andamento": True, "iniciado_em": _agora.isoformat(),
                  "por": current_user.get("name") or current_user.get("email", "")}},
        upsert=True,
    )

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        await db.import_locks.update_one({"chave": "aet"}, {"$set": {"em_andamento": False}})
        return {"ok": False, "message": f"Erro ao abrir Excel: {e}"}

    if "Analise" not in wb.sheetnames:
        await db.import_locks.update_one({"chave": "aet"}, {"$set": {"em_andamento": False}})
        return {"ok": False, "message": "Aba 'Analise' não encontrada no arquivo. Não parece ser a AET."}

    ws = wb["Analise"]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        await db.import_locks.update_one({"chave": "aet"}, {"$set": {"em_andamento": False}})
        return {"ok": False, "message": "Arquivo vazio"}

    # Pré-scan: conta quantas vezes cada numero_pedido aparece no arquivo
    # Se > 1, o pedido tem mais de um produto → is_parcial = True
    all_rows = list(rows_iter)
    contagem_pedidos = {}
    for row in all_rows:
        if not row or row[0] is None:
            continue
        entrega = str(row[0]).strip().split(".")[0]
        if entrega and entrega != "0":
            contagem_pedidos[entrega] = contagem_pedidos.get(entrega, 0) + 1
    rows_iter = iter(all_rows)  # reinicia iteração

    # Carrega os cancelamentos AES já existentes (o mais recente por numero_pedido)
    _CANCELADO = {"Cancelado", "Cancelada", "CANCELADO", "Entrega cancelada"}
    existentes = await db.cancelamentos.find(
        {"tipo": "aes"},
        {"numero_pedido": 1, "id": 1, "status": 1, "data_encerramento": 1, "observacao": 1, "_id": 0},
    ).sort("data_criacao", -1).to_list(None)
    aes_por_pedido = {}
    for c in existentes:
        np = c.get("numero_pedido")
        if np and np not in aes_por_pedido:
            aes_por_pedido[np] = c
    pedidos_existentes = set(aes_por_pedido.keys())
    reabertos = 0

    user_name = current_user.get("name") or current_user.get("email", "Sistema")
    now_iso = datetime.now(timezone.utc).isoformat()
    now_br = datetime.now(timezone(timedelta(hours=-3)))
    hoje_str = now_br.strftime("%Y-%m-%d")

    total_linhas = 0
    ignorados_decisao = 0
    ja_existiam = 0
    inseridos = 0
    enriquecidos = 0
    erros = 0
    sem_pedido = 0

    novos_docs = []

    for row in rows_iter:
        if not row or row[0] is None:
            continue
        total_linhas += 1
        try:
            entrega = str(row[0]).strip().split(".")[0]
            if not entrega or entrega == "0":
                sem_pedido += 1
                continue

            decisao_raw = row[19] if len(row) > 19 else None
            retorno_raw = str(row[18] or "") if len(row) > 18 else ""

            acao_norm = _classifica_decisao(decisao_raw)
            if not acao_norm:
                ignorados_decisao += 1
                continue

            is_parcial = contagem_pedidos.get(entrega, 1) > 1

            if entrega in pedidos_existentes:
                ex = aes_por_pedido[entrega]
                enc = str(ex.get("data_encerramento") or "")[:10]
                row_data = _to_date_iso(row[15] if len(row) > 15 else None) or hoje_str
                # Reabre se o AES está encerrado e a solicitação da AET é POSTERIOR ao
                # encerramento — exceto se o pedido já foi realmente cancelado.
                if ex.get("status") == "encerrado" and len(enc) >= 10 and str(row_data) > enc:
                    ped_c = await db.pedidos_erp.find_one({"numero_pedido": entrega}, {"_id": 0, "status_pedido": 1})
                    if ped_c and (ped_c.get("status_pedido") or "") in _CANCELADO:
                        ja_existiam += 1
                        continue
                    obs_ex = ex.get("observacao", "") or ""
                    nota = f"{hoje_str[8:10]}/{hoje_str[5:7]} - Reaberto: nova solicitação na planilha AET (após encerramento em {enc[8:10]}/{enc[5:7]})"
                    await db.cancelamentos.update_one(
                        {"id": ex["id"]},
                        {"$set": {"status": "pendente", "data_encerramento": "", "reaberto": True,
                                  "reaberto_em": now_iso, "reaberto_motivo": "planilha_aet",
                                  "observacao": (nota + ("\n" + obs_ex if obs_ex else "")).strip(),
                                  "updated_at": now_iso},
                         "$unset": {"encerrado_por_compras": "", "encerrado_por_etr": ""}},
                    )
                    reabertos += 1
                    continue
                ja_existiam += 1
                continue

            # Concluído (Sim/Não)
            concluido = _norm(row[21] if len(row) > 21 else "")
            status = "encerrado" if concluido == "sim" else "pendente"
            data_encerramento = hoje_str if status == "encerrado" else ""

            doc = {
                "id": str(uuid.uuid4()),
                "tipo": "aes",
                "status": status,
                "numero_pedido": entrega,
                "data_criacao": now_iso,
                "data": hoje_str,
                "data_status_aet": _to_date_iso(row[15] if len(row) > 15 else None),
                "criado_por": "Importação AET",
                "criado_por_email": current_user.get("email", ""),
                "motivo": str(row[17] or "")[:200] if len(row) > 17 else "",
                "acao": acao_norm,  # 'Cancelar' ou 'Similar'
                "is_parcial": is_parcial,  # True se o pedido tem mais de 1 produto na AET
                "motivo_rejeicao": "",
                "ticket": "",
                "instancia": "",
                "zerado_reserva": None,
                "observacao": retorno_raw[:500],
                "cliente_confirmado_nome": "",
                "cliente_confirmado_cpf": "",
                "cliente_confirmado_endereco": "",
                "nova_entrega": "",
                "parceiro_planilha": str(row[12] or "") if len(row) > 12 else "",
                "codigo_terceiro_planilha": str(row[4] or "") if len(row) > 4 else "",
                "filial_planilha": str(row[13] or "") if len(row) > 13 else "",
                "data_encerramento": data_encerramento,
                "aet_pedido_compra": str(row[11] or "") if len(row) > 11 else "",
                "aet_dias_parado": int(row[16]) if len(row) > 16 and row[16] is not None else None,
                "aet_decisao": str(decisao_raw or ""),
                "aet_status_atual": str(row[20] or "") if len(row) > 20 else "",
                "updated_at": now_iso,
            }

            # Enriquecer com tabelão. Em pedido MULTI-ITEM, o pedidos_erp costuma ter
            # só 1 dos N itens — então buscamos primeiro o ITEM da planilha (row[4]=SKU).
            # Antes, sem o SKU, pegava um item qualquer do pedido e sobrescrevia o SKU
            # correto vindo da AET (caso real: peneira virava caneca/tábua).
            sku_planilha = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else "")
            pedido = None
            if sku_planilha:
                pedido = await db.pedidos_erp.find_one({"numero_pedido": entrega, "codigo_item_vtex": sku_planilha})
            pedido_ordem = pedido or await db.pedidos_erp.find_one({"numero_pedido": entrega})
            if pedido_ordem:
                # Dados do PEDIDO (cliente/endereço/NF/transportadora) servem de qualquer item
                doc.update({
                    "canal_vendas": pedido_ordem.get("canal_vendas", ""),
                    "nome_cliente": pedido_ordem.get("nome_cliente", ""),
                    "cpf_cliente": pedido_ordem.get("cpf_cliente", ""),
                    "fone_cliente": pedido_ordem.get("fone_cliente", ""),
                    "email_cliente": pedido_ordem.get("email_cliente", ""),
                    "cep": pedido_ordem.get("cep", ""),
                    "cidade": pedido_ordem.get("cidade", ""),
                    "uf": pedido_ordem.get("uf", ""),
                    "endereco_rua": pedido_ordem.get("endereco_rua", ""),
                    "endereco_numero": pedido_ordem.get("endereco_numero", ""),
                    "endereco_complemento": pedido_ordem.get("endereco_complemento", ""),
                    "endereco_bairro": pedido_ordem.get("endereco_bairro", ""),
                    "filial": pedido_ordem.get("filial", ""),
                    "nota_fiscal": pedido_ordem.get("nota_fiscal", ""),
                    "serie_nf": pedido_ordem.get("serie_nf", ""),
                    "chave_nota": pedido_ordem.get("chave_nota", ""),
                    "transportadora": pedido_ordem.get("transportadora", ""),
                    "status_pedido": pedido_ordem.get("status_pedido", ""),
                    "data_status": pedido_ordem.get("data_status", ""),
                })
                # Dados do ITEM: usa o item CERTO (casado pelo SKU da planilha). Se o
                # tabelão não tiver esse item (multi-item), preserva o que veio da AET.
                item = pedido or {}
                doc["produto"] = item.get("produto") or (str(row[3] or "") if len(row) > 3 else "")
                doc["codigo_item_vtex"] = item.get("codigo_item_vtex") or sku_planilha
                doc["codigo_item_bseller"] = item.get("codigo_item_bseller", "")
                doc["codigo_fornecedor"] = item.get("codigo_fornecedor") or (str(row[5] or "") if len(row) > 5 else "")
                doc["departamento"] = item.get("departamento") or (str(row[6] or "") if len(row) > 6 else "")
                doc["preco_final"] = item.get("preco_final") or (str(row[8] or "") if len(row) > 8 else "")
                doc["quantidade"] = item.get("quantidade") or (str(row[7] or "") if len(row) > 7 else "")
                enriquecidos += 1
            else:
                doc["produto"] = str(row[3] or "") if len(row) > 3 else ""
                doc["codigo_item_vtex"] = str(row[4] or "") if len(row) > 4 else ""
                doc["codigo_fornecedor"] = str(row[5] or "") if len(row) > 5 else ""
                doc["departamento"] = str(row[6] or "") if len(row) > 6 else ""
                doc["preco_final"] = str(row[8] or "") if len(row) > 8 else ""
                doc["quantidade"] = str(row[7] or "") if len(row) > 7 else ""
                doc["canal_vendas"] = str(row[12] or "") if len(row) > 12 else ""

            novos_docs.append(doc)
            pedidos_existentes.add(entrega)
            inseridos += 1

        except Exception as e:
            logger.warning(f"Erro na linha AET: {e}")
            erros += 1
            continue

    # Inserir em lote — antes, RECHECA no banco quem já existe (cinto e suspensório
    # contra corrida: outra importação pode ter gravado durante o processamento).
    if novos_docs:
        ents_novos = list({d["numero_pedido"] for d in novos_docs})
        ja_no_banco = set()
        async for c in db.cancelamentos.find(
                {"tipo": "aes", "numero_pedido": {"$in": ents_novos}},
                {"_id": 0, "numero_pedido": 1}):
            ja_no_banco.add(c["numero_pedido"])
        if ja_no_banco:
            antes = len(novos_docs)
            novos_docs = [d for d in novos_docs if d["numero_pedido"] not in ja_no_banco]
            pulados = antes - len(novos_docs)
            inseridos -= pulados
            ja_existiam += pulados
            logger.warning(f"[AET] {pulados} doc(s) descartado(s) na rechecagem final (já existiam — corrida evitada)")
    if novos_docs:
        try:
            await db.cancelamentos.insert_many(novos_docs, ordered=False)
        except Exception as e:
            logger.error(f"Erro ao inserir cancelamentos AET: {e}")

    # Libera a trava de importação
    await db.import_locks.update_one({"chave": "aet"}, {"$set": {"em_andamento": False}})

    # Notificar admins
    try:
        partes = [f"{inseridos} novos cancelamentos AES inseridos"]
        if reabertos:
            partes.append(f"{reabertos} reaberto(s) por nova solicitação")
        partes.append(f"{ja_existiam} já existiam")
        msg = f"Base AET importada por {user_name}. {', '.join(partes)}."

        usuarios = await db.users.find({"role": {"$nin": ["system"]}}, {"email": 1, "_id": 0}).to_list(50)
        for u in usuarios:
            if not u.get("email"):
                continue
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "tipo": "import_concluida",
                "titulo": "Base AET Importada",
                "mensagem": msg,
                "destinatario_email": u["email"],
                "dados_extras": {
                    "inseridos": inseridos,
                    "ja_existiam": ja_existiam,
                    "ignorados_decisao": ignorados_decisao,
                    "enriquecidos": enriquecidos,
                    "erros": erros,
                },
                "data_criacao": now_iso,
                "lida": False,
                "criado_por_nome": "Sistema",
            })
    except Exception as e:
        logger.warning(f"Erro ao notificar AET: {e}")

    return {
        "ok": True,
        "total_linhas": total_linhas,
        "inseridos": inseridos,
        "reabertos": reabertos,
        "ja_existiam": ja_existiam,
        "ignorados_decisao": ignorados_decisao,
        "enriquecidos": enriquecidos,
        "erros": erros,
        "sem_pedido": sem_pedido,
        "message": f"{total_linhas} linhas processadas — {inseridos} novos, {reabertos} reaberto(s), {ja_existiam} já existiam.",
    }
