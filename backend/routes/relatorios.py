import io
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse

from utils.database import db
from utils.auth import get_current_user
from utils.helpers import calcular_dias_uteis

router = APIRouter(prefix="/api")

_BRT = timezone(timedelta(hours=-3))


def _gerar_relatorio_xlsx(titulo, headers, rows, header_hex, col_widths,
                          sheet_name, left_align,
                          header_font_hex="000000", band_hex="F2F2F2"):
    """Monta um .xlsx padronizado: título + 'Gerado em DD/MM/AAAA | Total: N',
    cabeçalho colorido com filtro, bordas, faixas alternadas e 'Crítico' em
    vermelho. `left_align` = set de cabeçalhos alinhados à esquerda."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ncols = len(headers)
    last_col = get_column_letter(ncols)
    data_str = datetime.now(_BRT).strftime("%d/%m/%Y")

    # Título (linha 1) + subtítulo (linha 2); linha 3 vazia; cabeçalho na 4
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = titulo
    t.font = Font(bold=True, size=14, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = f"Gerado em {data_str} | Total: {len(rows)}"
    s.font = Font(bold=True, size=11, color="595959", name="Calibri")
    s.alignment = Alignment(horizontal="center", vertical="center")

    HDR = 4
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=header_hex)
    header_font = Font(bold=True, color=header_font_hex, name="Calibri")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=HDR, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[HDR].height = 26

    band = PatternFill("solid", fgColor=band_hex)
    critico_font = Font(bold=True, color="C00000", name="Calibri")
    for i, row in enumerate(rows):
        r = HDR + 1 + i
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = border
            esq = headers[j - 1] in left_align
            c.alignment = Alignment(horizontal="left" if esq else "center", vertical="center")
            if i % 2 == 1:
                c.fill = band
            if isinstance(val, str) and val.strip() == "Crítico":
                c.font = critico_font

    for j, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.auto_filter.ref = f"A{HDR}:{last_col}{HDR + max(len(rows), 1)}"
    ws.freeze_panes = f"A{HDR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_response(buf, prefixo):
    fname = f"{prefixo}_{datetime.now(_BRT).strftime('%d-%m-%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/relatorios/ag-compras")
async def get_relatorio_ag_compras(current_user: dict = Depends(get_current_user)):
    # Pedidos que já estão na lista de cancelamento AES devem ser excluídos do relatório
    # (qualquer status: pendentes ainda em andamento ou encerrados já tratados)
    aes_pedidos = await db.cancelamentos.find(
        {"tipo": "aes"}, {"_id": 0, "numero_pedido": 1}
    ).to_list(5000)
    aes_numeros_pedido = {c["numero_pedido"] for c in aes_pedidos if c.get("numero_pedido")}

    chamados_query = {"motivo_pendencia": "Ag. Compras", "pendente": True}
    if aes_numeros_pedido:
        chamados_query["numero_pedido"] = {"$nin": list(aes_numeros_pedido)}

    chamados = await db.chamados.find(chamados_query, {"_id": 0}).to_list(5000)
    fornecedores_dict = {}
    fornecedores = await db.fornecedores.find({}, {"_id": 0}).to_list(100)
    for f in fornecedores:
        fornecedores_dict[f.get('nome', '').lower()] = f.get('dias_extras_padrao', 5)

    # Bulk query pedidos (evita N+1)
    pedido_numbers = [c.get('numero_pedido') for c in chamados if c.get('numero_pedido')]
    pedidos_raw = await db.pedidos_erp.find(
        {"numero_pedido": {"$in": pedido_numbers}}, {"_id": 0}
    ).to_list(len(pedido_numbers)) if pedido_numbers else []
    pedidos_dict = {p['numero_pedido']: p for p in pedidos_raw}

    # Estoque XD (cross-dock) = estoque_sigeq com source 'SIGEQ425', casado pelo SKU
    # (cod_terceiro = codigo_item_vtex). Item sem registro XD → 0 (nunca "-").
    skus = [str(p.get('codigo_item_vtex') or '').strip()
            for p in pedidos_dict.values() if p.get('codigo_item_vtex')]
    xd_raw = await db.estoque_sigeq.find(
        {"cod_terceiro": {"$in": skus}, "source": "SIGEQ425"},
        {"_id": 0, "cod_terceiro": 1, "disp_venda": 1, "codigo_fornecedor": 1}
    ).to_list(len(skus)) if skus else []
    xd_dict = {e['cod_terceiro']: e for e in xd_raw}

    resultado = []
    for chamado in chamados:
        pedido = pedidos_dict.get(chamado.get('numero_pedido'))
        if not pedido:
            continue
        status_pedido = pedido.get('status_pedido', '').lower()
        data_status = pedido.get('data_status', '')
        fornecedor = pedido.get('departamento', '')
        incluir = False
        if 'pedido aprovado' in status_pedido:
            incluir = True
        elif 'aguardando estoque' in status_pedido:
            dias_extras = fornecedores_dict.get(fornecedor.lower(), 5)
            dias_em_estoque = calcular_dias_uteis(data_status)
            if dias_em_estoque >= dias_extras:
                incluir = True
        if not incluir:
            continue

        status_atendimento = ""
        if chamado.get('retornar_chamado'):
            status_atendimento = "Retornar"
        elif chamado.get('verificar_adneia'):
            status_atendimento = "Verificar"

        sku = str(pedido.get('codigo_item_vtex') or '').strip()
        xd = xd_dict.get(sku)
        # Estoque XD: sempre número (0 quando não há registro), nunca None/"-"
        estoque_disponivel = int(xd.get('disp_venda') or 0) if xd else 0

        resultado.append({
            "fornecedor": fornecedor,
            "produto": pedido.get('produto', '') or chamado.get('produto', ''),
            "id_produto": pedido.get('codigo_item_bseller', ''),
            "sku": pedido.get('codigo_item_vtex', ''),
            "quantidade": pedido.get('quantidade', ''),
            "codigo_fornecedor": (xd.get('codigo_fornecedor') if xd and xd.get('codigo_fornecedor') else pedido.get('codigo_fornecedor', '')),
            "entrega": chamado.get('numero_pedido', ''),
            "parceiro_canal": pedido.get('canal_vendas', ''),
            "cidade": pedido.get('cidade', ''),
            "uf": pedido.get('uf', ''),
            "estoque_disponivel": estoque_disponivel,
            "status_atendimento": status_atendimento,
            "status_entrega": pedido.get('status_pedido', ''),
            "data_ultimo_ponto": data_status
        })
    return resultado


@router.get("/relatorios/ag-logistica")
async def get_relatorio_ag_logistica(current_user: dict = Depends(get_current_user)):
    chamados = await db.chamados.find(
        {"motivo_pendencia": "Ag. Logística", "pendente": True}, {"_id": 0}
    ).to_list(5000)

    # Bulk query pedidos (evita N+1)
    pedido_numbers = [c.get('numero_pedido') for c in chamados if c.get('numero_pedido')]
    pedidos_raw = await db.pedidos_erp.find(
        {"numero_pedido": {"$in": pedido_numbers}}, {"_id": 0}
    ).to_list(len(pedido_numbers)) if pedido_numbers else []
    pedidos_dict = {p['numero_pedido']: p for p in pedidos_raw}

    resultado = []
    for chamado in chamados:
        pedido = pedidos_dict.get(chamado.get('numero_pedido'))
        if not pedido:
            continue
        status_pedido = (pedido.get('status_pedido', '') or '').lower()
        data_status = pedido.get('data_status', '')
        # Regra (definida pela Adneia): SÓ pedidos em ETR (Entregue a Transportadora),
        # com MAIS DE 2 dias úteis nesse status. 'Pedido aprovado' e demais status NÃO entram.
        if 'entregue a transportadora' not in status_pedido:
            continue
        dias_no_status = calcular_dias_uteis(data_status)
        if dias_no_status <= 2:  # "mais de 2 dias úteis" = > 2 (inclui a partir de 3)
            continue

        status_atendimento = ""
        if chamado.get('retornar_chamado'):
            status_atendimento = "Retornar"
        elif chamado.get('verificar_adneia'):
            status_atendimento = "Verificar"

        resultado.append({
            "entrega": chamado.get('numero_pedido', ''),
            "nota": str(pedido.get('nota_fiscal', '')).replace('.0', '') if pedido.get('nota_fiscal') else '',
            "galpao": pedido.get('filial', ''),
            "status_entrega": pedido.get('status_pedido', ''),
            "data_ultimo_ponto": data_status,
            "dias_no_status": dias_no_status,
            "status_atendimento": status_atendimento
        })
    # Relatório em ORDEM DE GALPÃO (e, dentro de cada galpão, os mais parados primeiro)
    resultado.sort(key=lambda r: ((r.get('galpao') or '').upper(), -int(r.get('dias_no_status') or 0), str(r.get('entrega') or '')))
    return resultado


def _map_critico(status):
    return "Crítico" if status in ("Verificar", "Retornar") else (status or "")


@router.get("/relatorios/ag-compras-xlsx")
async def get_relatorio_ag_compras_xlsx(current_user: dict = Depends(get_current_user)):
    data = await get_relatorio_ag_compras(current_user)
    headers = ["Fornecedor", "Produto", "Cód. Fornecedor", "ID", "SKU", "Estoque XD",
               "Qtd. Pedido", "Entrega", "Parceiro/Canal", "Cidade", "UF",
               "Status Atendimento", "Status Entrega", "Data Último Ponto"]
    widths = [20, 38, 16, 12, 13, 11, 11, 13, 16, 18, 6, 16, 20, 20]
    rows = []
    for it in data:
        rows.append([
            it.get("fornecedor") or "",
            it.get("produto") or "",
            it.get("codigo_fornecedor") or "",
            it.get("id_produto") or "",
            it.get("sku") or "",
            it.get("estoque_disponivel") if it.get("estoque_disponivel") is not None else 0,
            it.get("quantidade") or "",
            it.get("entrega") or "",
            it.get("parceiro_canal") or "",
            it.get("cidade") or "",
            it.get("uf") or "",
            _map_critico(it.get("status_atendimento")),
            it.get("status_entrega") or "",
            it.get("data_ultimo_ponto") or "",
        ])
    buf = _gerar_relatorio_xlsx(
        "Atendimentos pendentes em Compras", headers, rows, "4BACC6", widths,
        "Ag Compras", {"Fornecedor", "Produto", "Cidade", "Status Entrega", "Parceiro/Canal"})
    return _xlsx_response(buf, "relatorio_ag_compras")


@router.get("/relatorios/ag-logistica-xlsx")
async def get_relatorio_ag_logistica_xlsx(current_user: dict = Depends(get_current_user)):
    data = await get_relatorio_ag_logistica(current_user)
    headers = ["Entrega", "Nota", "Galpão", "Status Entrega", "Data Último Ponto",
               "Dias em ETR", "Status Atendimento"]
    widths = [15, 15, 12, 24, 20, 12, 18]
    rows = []
    for it in data:
        rows.append([
            it.get("entrega") or "",
            it.get("nota") or "",
            it.get("galpao") or "",
            it.get("status_entrega") or "",
            it.get("data_ultimo_ponto") or "",
            it.get("dias_no_status") if it.get("dias_no_status") is not None else "",
            _map_critico(it.get("status_atendimento")),
        ])
    buf = _gerar_relatorio_xlsx(
        "Atendimentos pendentes na logística", headers, rows, "ED7D31", widths,
        "Ag Logistica", {"Status Entrega"})
    return _xlsx_response(buf, "relatorio_ag_logistica")


@router.post("/relatorios/acionar-parceiro-xlsx")
async def acionar_parceiro_xlsx(payload: dict, current_user: dict = Depends(get_current_user)):
    """Gera o xlsx estilizado do 'Acionar Parceiro' a partir das linhas já
    compiladas na tela (mesmo conteúdo do modal). Tema vermelho."""
    linhas = payload.get("linhas") or []
    if not linhas:
        raise HTTPException(status_code=400, detail="Sem linhas para exportar")
    headers = ["Parceiro", "Entrega", "Pedido", "Solicitação", "CPF",
               "Data Anotação", "Última Anotação"]
    widths = [16, 15, 18, 14, 16, 14, 62]
    rows = [[
        l.get("parceiro") or "",
        l.get("entrega") or "",
        l.get("pedido") or "",
        l.get("solicitacao") or "",
        l.get("cpf") or "",
        l.get("data_anotacao") or "",
        l.get("anotacao") or "",
    ] for l in linhas]
    parceiros = sorted({(l.get("parceiro") or "").strip() for l in linhas
                        if (l.get("parceiro") or "").strip() not in ("", "-")})
    rotulo = " + ".join(parceiros) if 0 < len(parceiros) <= 3 else ("vários" if parceiros else "")
    titulo = f"Base Acionar Parceiro{f' - {rotulo}' if rotulo else ''}"
    buf = _gerar_relatorio_xlsx(
        titulo, headers, rows, "C0504D", widths, "Acionar Parceiro",
        {"Parceiro", "Pedido", "Última Anotação"},
        header_font_hex="FFFFFF", band_hex="F2DCDB")
    return _xlsx_response(buf, "acionar_parceiro")
