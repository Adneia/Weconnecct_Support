from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, BackgroundTasks
from typing import Optional, List
from datetime import datetime, timezone, timedelta
import os
import re
import uuid

import httpx

from utils.database import db
from utils.auth import get_current_user
from utils.helpers import (
    parse_date_safe, get_galpao_from_serie, get_column_mapping,
    extract_pedido_data, should_skip_old_pedido
)

import logging
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")

# ====== BSeller SAC API (rastreio em tempo real) ======
BSELLER_SAC_URL = "https://api.bseller.com.br/sac/atendimento/entregas"
BSELLER_TOKEN = os.getenv("BSELLER_TOKEN", "")

# Prefixos de usuario classificados como automacao no rastreio SAC
# (resto = humano). Ver bseller-api-map/exports/sac_usuario_padroes.csv
_USR_AUTOMACAO = ("INTELIPOST", "FAT_AUTO", "WEBSERVICE", "TUDOAZUL", "SUPERTROCO", "ROOT")


def _classify_usuario(usuario: str | None) -> str:
    if not usuario:
        return "unknown"
    u = usuario.upper()
    if any(k in u for k in _USR_AUTOMACAO):
        return "automacao"
    if u.startswith("FUNS_"):
        return "humano"
    return "unknown"


def _format_endereco(end: dict) -> str:
    """Concatena endereco em string amigavel pro UI do ELO."""
    if not end:
        return ""
    parts = []
    logradouro = (end.get("logradouro") or "").strip()
    numero = end.get("numero")
    complemento = (end.get("complemento") or "").strip()
    bairro = (end.get("bairro") or "").strip()
    cidade = (end.get("cidade") or "").strip()
    estado = (end.get("estado") or "").strip()
    cep = (end.get("cep") or "").strip()

    if logradouro:
        if numero is not None:
            parts.append(f"{logradouro}, {numero}")
        else:
            parts.append(logradouro)
    if complemento:
        parts.append(complemento)
    if bairro:
        parts.append(bairro)
    cidade_uf = ", ".join(p for p in (cidade, estado) if p)
    if cidade_uf:
        parts.append(cidade_uf)
    if cep:
        parts.append(f"CEP {cep}")
    return " — ".join(parts)


async def _enrich_pedido_endereco_via_sac(pedido: dict) -> dict:
    """Enriquece endereco do pedido com dados da API SAC do BSeller quando incompleto.

    QRY0010 nao traz logradouro/numero/complemento/bairro do destino, so CEP+cidade+UF.
    O upload do Tabelao Excel preenche rua/numero/bairro pra maioria dos pedidos (99%),
    mas em ~37% deles cep e cidade ficam vazios (caso onde cliente cadastrado em cidade
    diferente da entrega).

    Este helper chama GET /sac/atendimento/entregas/{pedidoCliente} sob demanda,
    preenche os campos faltantes e persiste no Mongo pra proxima leitura ser local.

    Fail-safe: se BSELLER_TOKEN nao configurado, ID nao numerico, ou API SAC falhar,
    retorna o pedido original sem erro.
    """
    if not BSELLER_TOKEN:
        return pedido

    # Detecta se endereco ja esta completo (campos minimos)
    has_rua = bool((pedido.get('endereco_rua') or '').strip())
    has_cep = bool((pedido.get('cep') or '').strip())
    has_cidade = bool((pedido.get('cidade') or '').strip())
    if has_rua and has_cep and has_cidade:
        return pedido

    # API SAC do BSeller exige o pedido_bseller (LONG numerico ~700-900k).
    # No ELO Mongo TEST o mapeamento eh:
    #   pedido_cliente = pedido_bseller real (6 digitos, ~719k)  <- API SAC aceita
    #   numero_pedido  = id_entrega         (9 digitos, ~122M)   <- API SAC retorna 400
    #   codigo_pedido  = quase sempre vazio
    # Tentamos cada candidato em ordem ate achar um HTTP 200.
    candidates_raw = [
        pedido.get('pedido_cliente'),   # mais provavel: pedido_bseller real
        pedido.get('codigo_pedido'),    # fallback
        pedido.get('numero_pedido'),    # fallback (pode acertar em uploads diferentes)
        pedido.get('pedido_externo'),
    ]
    candidates = []
    seen = set()
    for c in candidates_raw:
        if not c:
            continue
        s = str(c).strip()
        if re.match(r'^[0-9]+$', s) and s not in seen:
            candidates.append(s)
            seen.add(s)
    if not candidates:
        return pedido

    headers = {"x-auth-token": BSELLER_TOKEN, "content-type": "application/json"}
    data = None
    pedido_bseller_used = None
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            for cand in candidates:
                resp = await client.get(f"{BSELLER_SAC_URL}/{cand}", headers=headers)
                if resp.status_code == 200:
                    data = resp.json()
                    pedido_bseller_used = cand
                    break
                # 400/404 = pedido nao corresponde a este ID — tenta o proximo
    except Exception as exc:
        logger.warning(f"SAC enrich falhou pra {pedido.get('numero_pedido')}: {exc}")
        return pedido

    if not data:
        return pedido

    entregas = data.get('entregas') or []
    if not entregas:
        return pedido

    # Pega a entrega que tem logradouro (caso multi-entrega). Fallback: primeira.
    ent_with_addr = next(
        (e for e in entregas if ((e.get('clienteEntrega') or {}).get('endereco') or {}).get('logradouro')),
        entregas[0],
    )
    end = (ent_with_addr.get('clienteEntrega') or {}).get('endereco') or {}

    # So preenche campos faltantes (nao sobrescreve o que veio do Tabelao)
    update_fields = {}
    if not has_rua and end.get('logradouro'):
        update_fields['endereco_rua'] = end.get('logradouro')
    if (not (pedido.get('endereco_numero') or '').strip()) and end.get('numero') is not None:
        update_fields['endereco_numero'] = str(end.get('numero'))
    if (not (pedido.get('endereco_complemento') or '').strip()) and end.get('complemento'):
        update_fields['endereco_complemento'] = end.get('complemento')
    if (not (pedido.get('endereco_bairro') or '').strip()) and end.get('bairro'):
        update_fields['endereco_bairro'] = end.get('bairro')
    if not has_cep and end.get('cep'):
        update_fields['cep'] = end.get('cep')
    if not has_cidade and end.get('cidade'):
        update_fields['cidade'] = end.get('cidade')
    if (not (pedido.get('uf') or '').strip()) and end.get('estado'):
        update_fields['uf'] = end.get('estado')

    if not update_fields:
        return pedido

    update_fields['endereco_sac_synced_at'] = datetime.now(timezone.utc).isoformat()
    update_fields['endereco_sac_source'] = 'bseller_sac_atendimento_entregas'
    update_fields['endereco_sac_pedido_bseller'] = pedido_bseller_used

    # Persiste no Mongo pra proxima leitura ser local (zero hit em API SAC)
    try:
        await db.pedidos_erp.update_one(
            {"numero_pedido": pedido.get('numero_pedido')},
            {"$set": update_fields},
        )
    except Exception as exc:
        logger.warning(f"SAC enrich: falha persistindo {pedido.get('numero_pedido')}: {exc}")

    # Aplica no objeto retornado tambem
    pedido.update(update_fields)
    logger.info(
        f"SAC enrich aplicado em {pedido.get('numero_pedido')}: "
        f"{list(update_fields.keys())}"
    )
    return pedido


# ============== BUSCAR PEDIDOS ==============

@router.get("/pedidos-erp/buscar/cpf/{cpf}")
async def buscar_pedido_por_cpf(cpf: str, current_user: dict = Depends(get_current_user)):
    """Search pedidos by CPF (supports both formatted and unformatted)"""
    # Remove formatting from CPF
    cpf_limpo = cpf.replace(".", "").replace("-", "").strip()
    
    pedidos = await db.pedidos_erp.find({
        "$or": [
            {"cpf_cliente": {"$regex": cpf_limpo, "$options": "i"}},
            {"cpf_cliente": {"$regex": cpf, "$options": "i"}}
        ]
    }, {"_id": 0}).sort("data_status", -1).to_list(50)
    
    # Add galpao info to each pedido
    for p in pedidos:
        galpao_info = get_galpao_from_serie(p.get('serie_nf', ''), p.get('chave_nota', ''))
        p['galpao'] = galpao_info.get('galpao', '')
        p['uf_galpao'] = galpao_info.get('uf_galpao', '')
    
    return pedidos


@router.get("/pedidos-erp/buscar/nome/{nome}")
async def buscar_pedido_por_nome(nome: str, current_user: dict = Depends(get_current_user)):
    """Search pedidos by customer name"""
    pedidos = await db.pedidos_erp.find({
        "nome_cliente": {"$regex": nome, "$options": "i"}
    }, {"_id": 0}).sort("data_status", -1).to_list(50)
    
    for p in pedidos:
        galpao_info = get_galpao_from_serie(p.get('serie_nf', ''), p.get('chave_nota', ''))
        p['galpao'] = galpao_info.get('galpao', '')
        p['uf_galpao'] = galpao_info.get('uf_galpao', '')
    
    return pedidos


@router.get("/pedidos-erp/buscar/pedido/{pedido}")
async def buscar_pedido_por_numero_pedido(pedido: str, current_user: dict = Depends(get_current_user)):
    """Search pedidos by any pedido-related field (partial match)"""
    pedidos = await db.pedidos_erp.find({
        "$or": [
            {"numero_pedido": {"$regex": pedido, "$options": "i"}},
            {"codigo_pedido": {"$regex": pedido, "$options": "i"}},
            {"pedido_cliente": {"$regex": pedido, "$options": "i"}},
            {"pedido_externo": {"$regex": pedido, "$options": "i"}}
        ]
    }, {"_id": 0}).sort("data_status", -1).to_list(50)
    
    for p in pedidos:
        galpao_info = get_galpao_from_serie(p.get('serie_nf', ''), p.get('chave_nota', ''))
        p['galpao'] = galpao_info.get('galpao', '')
        p['uf_galpao'] = galpao_info.get('uf_galpao', '')
    
    return pedidos


@router.get("/pedidos-erp/buscar/galpao/{galpao}/nota/{nota}")
async def buscar_pedido_por_galpao_nota(galpao: str, nota: str, current_user: dict = Depends(get_current_user)):
    """Search pedidos by galpao and nota fiscal"""
    query = {}
    nota_str = nota.strip()
    
    # Map galpao to serie_nf
    if galpao.upper() == "SC":
        query["serie_nf"] = "1"
    elif galpao.upper() == "SP":
        query["serie_nf"] = "6"
    elif galpao.upper() == "ES":
        query["serie_nf"] = "2"
    else:
        query["serie_nf"] = galpao
    
    query["$or"] = [
        {"nota_fiscal": nota_str},
        {"nota_fiscal": nota_str + ".0"},
        {"nota_fiscal": {"$regex": f"^{nota_str}", "$options": "i"}}
    ]
    
    pedidos = await db.pedidos_erp.find(query, {"_id": 0}).to_list(50)
    
    # If no results with serie_nf filter, try just nota
    if not pedidos:
        query_nota = {"$or": [
            {"nota_fiscal": nota_str},
            {"nota_fiscal": nota_str + ".0"},
            {"nota_fiscal": {"$regex": f"^{nota_str}", "$options": "i"}}
        ]}
        pedidos = await db.pedidos_erp.find(query_nota, {"_id": 0}).to_list(50)
    
    for p in pedidos:
        galpao_info = get_galpao_from_serie(p.get('serie_nf', ''), p.get('chave_nota', ''))
        p['galpao'] = galpao_info.get('galpao', '')
        p['uf_galpao'] = galpao_info.get('uf_galpao', '')
    
    return pedidos


@router.get("/pedidos-erp/import-status")
async def get_import_status(current_user: dict = Depends(get_current_user)):
    status = await db.import_status.find({}, {"_id": 0}).sort("started_at", -1).to_list(5)
    return status


@router.get("/pedidos-erp/import-status/{import_id}")
async def get_import_status_by_id(import_id: str, current_user: dict = Depends(get_current_user)):
    status = await db.import_status.find_one({"import_id": import_id}, {"_id": 0})
    if not status:
        return {"import_id": import_id, "status": "processing", "progress": 0, "total_rows": 0, "processed": 0}
    return status


@router.get("/pedidos-erp/{numero_pedido}")
async def get_pedido_by_entrega(numero_pedido: str, current_user: dict = Depends(get_current_user)):
    """Get single pedido by numero_pedido, codigo_pedido, pedido_cliente ou pedido_externo"""
    pedido = await db.pedidos_erp.find_one({"numero_pedido": numero_pedido}, {"_id": 0})
    
    # Buscar em campos alternativos se não encontrou
    if not pedido:
        pedido = await db.pedidos_erp.find_one({"codigo_pedido": numero_pedido}, {"_id": 0})
    if not pedido:
        pedido = await db.pedidos_erp.find_one({"pedido_cliente": numero_pedido}, {"_id": 0})
    if not pedido:
        pedido = await db.pedidos_erp.find_one({"pedido_externo": numero_pedido}, {"_id": 0})
    
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    
    # Add galpao info
    galpao_info = get_galpao_from_serie(pedido.get('serie_nf', ''), pedido.get('chave_nota', ''))
    pedido['galpao'] = galpao_info.get('galpao', '')
    pedido['uf_galpao'] = galpao_info.get('uf_galpao', '')

    # Enriquecer com código do fornecedor do SIGEQ
    id_item = pedido.get('codigo_item_bseller')
    if id_item:
        sigeq = await db.estoque_sigeq.find_one({"id_item": str(id_item)}, {"codigo_fornecedor": 1, "_id": 0})
        if sigeq and sigeq.get('codigo_fornecedor'):
            pedido['codigo_fornecedor'] = sigeq['codigo_fornecedor']

    # Enriquece endereco via API SAC quando incompleto (fail-safe)
    pedido = await _enrich_pedido_endereco_via_sac(pedido)

    return pedido


@router.get("/pedidos-erp/{numero_pedido}/rastreio-realtime")
async def get_rastreio_realtime(
    numero_pedido: str,
    current_user: dict = Depends(get_current_user),
):
    """Consulta GET /sac/atendimento/entregas do BSeller pra retornar:
    - status corrente (idPonto + descricao + dataPonto)
    - usuario que moveu (com tipo: humano/automacao)
    - endereco completo do cliente (logradouro, numero, complemento, bairro, cidade, UF, CEP)
    - itens e classificacao SAC quando disponivel

    Resolve a P0 do endereço (100% cobertura via API SAC) e da status real-time (vs BD com lag).

    Mapeamento completo: bseller-api-map/docs/rest/instancias_sac.md
    """
    if not BSELLER_TOKEN:
        raise HTTPException(
            status_code=503,
            detail="BSELLER_TOKEN nao configurado no backend ELO",
        )

    # A API SAC exige pedido_bseller (LONG ~700-900k). No ELO Mongo TEST o frontend
    # tipicamente passa numero_pedido (= id_entrega, ~120M) que NAO eh aceito.
    # Estrategia: tenta o input direto e, se nao bater, busca no Mongo outros candidatos
    # (pedido_cliente, codigo_pedido, pedido_externo) e tenta cada um.
    candidates = []
    seen = set()
    if re.match(r"^[0-9]+$", str(numero_pedido).strip()):
        candidates.append(str(numero_pedido).strip())
        seen.add(str(numero_pedido).strip())

    # Busca o pedido no Mongo pra coletar candidatos alternativos
    try:
        pedido_mongo = await db.pedidos_erp.find_one(
            {"$or": [
                {"numero_pedido": numero_pedido},
                {"codigo_pedido": numero_pedido},
                {"pedido_cliente": numero_pedido},
                {"pedido_externo": numero_pedido},
            ]},
            {"_id": 0, "numero_pedido": 1, "codigo_pedido": 1, "pedido_cliente": 1, "pedido_externo": 1},
        )
    except Exception:
        pedido_mongo = None

    if pedido_mongo:
        for key in ("pedido_cliente", "codigo_pedido", "numero_pedido", "pedido_externo"):
            v = pedido_mongo.get(key)
            if not v:
                continue
            s = str(v).strip()
            if re.match(r"^[0-9]+$", s) and s not in seen:
                candidates.append(s)
                seen.add(s)

    if not candidates:
        return {
            "status": "id_invalido",
            "pedido_bseller": numero_pedido,
            "mensagem": "API SAC do BSeller exige ID numerico (sem sufixo). Este pedido nao pode ser consultado.",
            "entregas": [],
        }

    headers = {"x-auth-token": BSELLER_TOKEN, "content-type": "application/json"}
    resp = None
    pedido_bseller_used = None
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            for cand in candidates:
                resp = await client.get(f"{BSELLER_SAC_URL}/{cand}", headers=headers)
                if resp.status_code == 200:
                    pedido_bseller_used = cand
                    break
                # 400/404 = ID candidato nao corresponde — tenta o proximo
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Timeout consultando BSeller")
    except httpx.RequestError as exc:
        logger.warning(f"Erro consultando SAC: {exc}")
        raise HTTPException(status_code=502, detail=f"Erro consultando BSeller: {exc}")

    if not pedido_bseller_used or resp is None or resp.status_code != 200:
        last_code = resp.status_code if resp is not None else "?"
        return {
            "status": "nao_encontrado",
            "pedido_bseller": numero_pedido,
            "candidatos_tentados": candidates,
            "mensagem": f"Nenhum candidato achou o pedido no BSeller (ultimo HTTP {last_code}).",
            "entregas": [],
        }

    try:
        data = resp.json()
    except Exception:
        raise HTTPException(status_code=502, detail="BSeller retornou JSON invalido")

    entregas_raw = data.get("entregas") or []
    entregas = []
    for ent in entregas_raw:
        rast = ent.get("rastreio") or {}
        cliente = ent.get("clienteEntrega") or {}
        end = cliente.get("endereco") or {}

        usuario = (rast.get("usuario") or "").strip() or None

        entregas.append({
            "id_entrega": str(ent.get("idEntrega") or ""),
            "id_filial": ent.get("idFilial"),
            "filial_uf": {2: "ES", 3: "SP", 4: "SC"}.get(ent.get("idFilial")),

            "id_ponto": rast.get("idPonto"),
            "ponto_descricao": rast.get("descricao"),
            "data_ponto": rast.get("dataPonto"),
            "usuario": usuario,
            "usuario_tipo": _classify_usuario(usuario),

            "cliente": {
                "id": str(cliente.get("id") or "") or None,
                "nome": cliente.get("nome"),
            },
            "endereco": {
                "logradouro": end.get("logradouro"),
                "numero": end.get("numero"),
                "complemento": end.get("complemento"),
                "bairro": end.get("bairro"),
                "cidade": end.get("cidade"),
                "estado": end.get("estado"),
                "cep": end.get("cep"),
                "pais": end.get("pais"),
                "ponto_referencia": end.get("pontoReferencia"),
                "completo": _format_endereco(end),
            },

            "itens": ent.get("itens") or [],
            "matriz_classificacao": [
                m for m in (ent.get("matrizClassificacao") or [])
                if any(v is not None for v in m.values())
            ],

            "id_meio_pagamento_principal": ent.get("idMeioPagamentoPrincipal"),
        })

    return {
        "status": "ok",
        "pedido_bseller": pedido_bseller_used,
        "numero_pedido_consultado": numero_pedido,
        "quantidade_entregas": data.get("quantidadeEntregas") or len(entregas),
        "consultado_em": datetime.now(timezone.utc).isoformat(),
        "entregas": entregas,
    }


@router.get("/pedidos-erp/{numero_pedido}/historico-rastreio")
async def get_historico_rastreio(
    numero_pedido: str,
    current_user: dict = Depends(get_current_user),
):
    """Retorna a linha do tempo completa de eventos de rastreio do pedido.

    Lê tracking_eventos no Postgres BIG-DATA (alimentado pelo relatório
    ZBIQ0035 do BSeller, refresh 6/6h). Quando o pedido tem multiplas
    entregas (irmãs no mesmo pedido_bseller), retorna a timeline de cada
    uma — exatamente o caso de pedidos multi-entrega que confundem o
    QRY0010.

    Estrutura:
        {
          "status": "ok",
          "pedido_bseller_real": "790117",
          "numero_pedido_consultado": "122703670",
          "quantidade_entregas": 2,
          "entregas": [
            {
              "id_entrega": "122703670",
              "total_eventos": 8,
              "eventos": [
                {"ponto_id": "PEI", "descricao": "...", "data_ocorrencia": "...",
                 "usuario": "FUNS_ADNEIA", "usuario_tipo": "humano", "source_api": "ZBIQ0035"},
                ...ordenados por data_ocorrencia ASC
              ]
            },
            ...
          ]
        }
    """
    import os
    import psycopg2
    from psycopg2.extras import RealDictCursor

    pg_host = os.getenv("PG_HOST")
    if not pg_host:
        raise HTTPException(
            status_code=503,
            detail="PG_HOST nao configurado no backend (sem conexao BIG-DATA)",
        )

    try:
        conn = psycopg2.connect(
            host=pg_host,
            port=os.getenv("PG_PORT", "5432"),
            dbname=os.getenv("PG_DB", "bigdata"),
            user=os.getenv("PG_USER"),
            password=os.getenv("PG_PASSWORD"),
            connect_timeout=5,
        )
    except Exception as exc:
        logger.warning(f"PG connect falhou: {exc}")
        raise HTTPException(status_code=502, detail=f"Sem conexao com BIG-DATA: {exc}")

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # 1) Resolve o pedido_bseller real a partir do id_entrega informado
            cur.execute(
                "SELECT pedido_bseller FROM pedidos WHERE id_entrega = %s LIMIT 1",
                (numero_pedido,),
            )
            row = cur.fetchone()
            if not row:
                # Talvez o usuário tenha passado o pedido_bseller diretamente
                cur.execute(
                    "SELECT DISTINCT pedido_bseller FROM pedidos WHERE pedido_bseller = %s LIMIT 1",
                    (numero_pedido,),
                )
                row = cur.fetchone()
                if not row:
                    return {
                        "status": "nao_encontrado",
                        "numero_pedido_consultado": numero_pedido,
                        "mensagem": "Pedido nao encontrado no BIG-DATA",
                        "entregas": [],
                    }

            pedido_bseller = row["pedido_bseller"]

            # 2) Lista todas as id_entrega desse pedido_bseller (multi-entrega)
            cur.execute(
                "SELECT DISTINCT id_entrega FROM pedidos WHERE pedido_bseller = %s "
                "AND id_entrega IS NOT NULL AND id_entrega <> '' ORDER BY id_entrega",
                (pedido_bseller,),
            )
            ids = [r["id_entrega"] for r in cur.fetchall()]

            # 3) Pra cada id_entrega, busca eventos no tracking_eventos
            # OBS: tracking_eventos.pedido_bseller na realidade armazena o id_entrega
            #      (nome legado da coluna). Ver feedback_pedidos_unique_reentrega.md
            entregas = []
            for id_ent in ids:
                cur.execute(
                    "SELECT ponto_id, descricao, data_ocorrencia, usuario, source_api "
                    "FROM tracking_eventos "
                    "WHERE pedido_bseller = %s "
                    "ORDER BY data_ocorrencia ASC, id ASC",
                    (id_ent,),
                )
                eventos_raw = cur.fetchall()
                eventos = [
                    {
                        "ponto_id": e["ponto_id"],
                        "descricao": e["descricao"],
                        "data_ocorrencia": e["data_ocorrencia"].isoformat() if e["data_ocorrencia"] else None,
                        "usuario": e["usuario"],
                        "usuario_tipo": _classify_usuario(e["usuario"]),
                        "source_api": e["source_api"],
                    }
                    for e in eventos_raw
                ]
                entregas.append({
                    "id_entrega": id_ent,
                    "total_eventos": len(eventos),
                    "ultimo_evento": eventos[-1] if eventos else None,
                    "eventos": eventos,
                })

        return {
            "status": "ok",
            "pedido_bseller_real": pedido_bseller,
            "numero_pedido_consultado": numero_pedido,
            "quantidade_entregas": len(entregas),
            "entregas": entregas,
        }
    finally:
        conn.close()


@router.get("/pedidos-erp/buscar")
async def buscar_pedido_erp(
    numero_pedido: Optional[str] = None,
    galpao: Optional[str] = None,
    nota: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if galpao and nota:
        query = {}
        nota_str = nota.strip()
        if galpao.upper() == "SC":
            query["serie_nf"] = "1"
        elif galpao.upper() == "SP":
            query["serie_nf"] = "6"
        elif galpao.upper() == "ES":
            query["serie_nf"] = "2"
        else:
            query["serie_nf"] = galpao
        query["$or"] = [
            {"nota_fiscal": nota_str},
            {"nota_fiscal": nota_str + ".0"},
        ]
        pedidos = await db.pedidos_erp.find(query, {"_id": 0}).to_list(10)
        if not pedidos:
            query_nota = {"$or": [
                {"nota_fiscal": nota_str},
                {"nota_fiscal": nota_str + ".0"},
            ]}
            pedidos = await db.pedidos_erp.find(query_nota, {"_id": 0}).to_list(10)
        if pedidos:
            for p in pedidos:
                galpao_info = get_galpao_from_serie(p.get('serie_nf', ''), p.get('chave_nota', ''))
                p['galpao'] = galpao_info.get('galpao', '')
                p['uf_galpao'] = galpao_info.get('uf_galpao', '')
                id_item = p.get('codigo_item_bseller')
                if id_item:
                    sigeq = await db.estoque_sigeq.find_one({"id_item": str(id_item)}, {"codigo_fornecedor": 1, "_id": 0})
                    if sigeq and sigeq.get('codigo_fornecedor'):
                        p['codigo_fornecedor'] = sigeq['codigo_fornecedor']
            return pedidos
        return []

    if not numero_pedido:
        raise HTTPException(status_code=400, detail="Número do pedido é obrigatório")

    pedido = await db.pedidos_erp.find_one({"numero_pedido": numero_pedido}, {"_id": 0})
    if not pedido:
        return []

    galpao_info = get_galpao_from_serie(pedido.get('serie_nf', ''), pedido.get('chave_nota', ''))
    pedido['galpao'] = galpao_info.get('galpao', '')
    pedido['uf_galpao'] = galpao_info.get('uf_galpao', '')
    id_item = pedido.get('codigo_item_bseller')
    if id_item:
        sigeq = await db.estoque_sigeq.find_one({"id_item": str(id_item)}, {"codigo_fornecedor": 1, "_id": 0})
        if sigeq and sigeq.get('codigo_fornecedor'):
            pedido['codigo_fornecedor'] = sigeq['codigo_fornecedor']
    return [pedido]


@router.get("/pedidos-erp")
async def list_pedidos_erp(
    page: int = 1,
    page_size: int = 50,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"numero_pedido": search_regex},
            {"cpf_cliente": search_regex},
            {"nome_cliente": search_regex},
            {"produto": search_regex}
        ]
    skip = (page - 1) * page_size
    total = await db.pedidos_erp.count_documents(query)
    pedidos = await db.pedidos_erp.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    return {"total": total, "page": page, "page_size": page_size, "pedidos": pedidos}


# ============== IMPORTAR PEDIDOS ==============

async def process_import_background(content: bytes, filename: str, user_name: str, user_email: str, import_id: str = None):
    import pandas as pd
    from io import BytesIO
    from routes.admin import atualizar_motivos_pendencia_automatico

    try:
        if not import_id:
            import_id = str(uuid.uuid4())[:8]
        await db.import_status.insert_one({
            "import_id": import_id,
            "status": "processing",
            "progress": 0,
            "total": 0,
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "errors": 0,
            "started_at": datetime.now(timezone.utc).isoformat(),
            "started_by": user_name
        })

        if filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        else:
            excel_file = pd.ExcelFile(BytesIO(content))
            sheet_names = excel_file.sheet_names
            logger.info(f"Abas encontradas: {sheet_names}")

            # Detectar arquivo "outras" (Fornecedores + Estoque, sem Tabelão)
            is_outras = 'Fornecedores' in sheet_names and 'Tabelão' not in sheet_names
            if is_outras:
                logger.info("Arquivo 'outras' detectado — importando Fornecedores e Estoque")
                # Importar Fornecedores
                if 'Fornecedores' in sheet_names:
                    df_forn = pd.read_excel(excel_file, sheet_name='Fornecedores')
                    df_forn.columns = df_forn.columns.str.strip().str.lower()
                    forn_count = 0
                    for _, row in df_forn.iterrows():
                        fornecedor = None
                        for col in ['fornecedor', 'nome', 'nome_fornecedor']:
                            if col in df_forn.columns:
                                fornecedor = str(row.get(col, '')).strip()
                                break
                        dias_extras = 5
                        for col in ['dias extras padrão (dias úteis)', 'dias extras padrão', 'dias extras', 'dias_extras']:
                            if col in df_forn.columns:
                                val = row.get(col)
                                if pd.notna(val):
                                    try: dias_extras = int(val)
                                    except: pass
                                break
                        if fornecedor and fornecedor.lower() not in ('nan', ''):
                            await db.fornecedores.update_one(
                                {"nome": fornecedor},
                                {"$set": {"nome": fornecedor, "dias_extras_padrao": dias_extras, "ultima_atualizacao": datetime.now(timezone.utc).isoformat()}},
                                upsert=True
                            )
                            forn_count += 1
                    logger.info(f"Fornecedores importados: {forn_count}")

                # Importar Estoque SIGEQ425 e SIGEQ230 com bulk_write (muito mais rápido)
                async def import_estoque_sheet(df_est, sheet, imp_id, progress_base):
                    from pymongo import UpdateOne
                    df_est.columns = df_est.columns.str.strip()
                    agora = datetime.now(timezone.utc).isoformat()
                    ops = []
                    total = len(df_est)
                    for _, row in df_est.iterrows():
                        id_item = str(row.get('ID do item', '')).strip()
                        if not id_item or id_item == 'nan': continue
                        if id_item.endswith('.0'): id_item = id_item[:-2]
                        data = {
                            "id_item": id_item,
                            "fornecedor": str(row.get('Nome do fornecedor', '')).strip(),
                            "descricao": str(row.get('Descrição do item', '')).strip(),
                            "codigo_fornecedor": str(row.get('Código fornecedor', '')).strip(),
                            "qt_reserva": int(row.get('Qt. Res', 0)) if pd.notna(row.get('Qt. Res')) else 0,
                            "disp_venda": int(row.get('Disp. Venda', 0)) if pd.notna(row.get('Disp. Venda')) else 0,
                            "qt_arquivo": int(row.get('Qt. Arquivo', 0)) if pd.notna(row.get('Qt. Arquivo')) else 0,
                            "sheet": sheet,
                            "ultima_atualizacao": agora
                        }
                        ops.append(UpdateOne({"id_item": id_item}, {"$set": data}, upsert=True))
                    # Executar em lotes de 1000
                    imp = upd = 0
                    batch_size = 1000
                    for i in range(0, len(ops), batch_size):
                        batch = ops[i:i+batch_size]
                        result = await db.estoque_sigeq.bulk_write(batch, ordered=False)
                        imp += result.upserted_count
                        upd += result.modified_count
                        progress = progress_base + int((min(i + batch_size, len(ops)) / max(len(ops), 1)) * 45)
                        await db.import_status.update_one(
                            {"import_id": imp_id},
                            {"$set": {"progress": progress, "inserted": imp, "updated": upd, "total": total}}
                        )
                    logger.info(f"Estoque {sheet}: {imp} novos, {upd} atualizados")
                    return imp, upd

                est_inserted = est_updated = 0
                sheets_to_process = [s for s in ['SIGEQ425', 'SIGEQ230'] if s in sheet_names]
                for idx_sheet, sheet in enumerate(sheets_to_process):
                    df_est = pd.read_excel(excel_file, sheet_name=sheet)
                    progress_base = int((idx_sheet / max(len(sheets_to_process), 1)) * 90)
                    i, u = await import_estoque_sheet(df_est, sheet, import_id, progress_base)
                    est_inserted += i; est_updated += u

                await db.import_status.update_one(
                    {"import_id": import_id},
                    {"$set": {"status": "completed", "progress": 100, "inserted": est_inserted, "updated": est_updated, "skipped": 0, "errors": 0, "total": est_inserted + est_updated, "completed_at": datetime.now(timezone.utc).isoformat()}}
                )
                logger.info(f"Import 'outras' concluído: {forn_count} fornecedores, {est_inserted} estoque novos, {est_updated} atualizados")
                # Notificar usuário que iniciou a importação
                try:
                    user = await db.users.find_one({"email": user_email}, {"_id": 0})
                    if not user:
                        user = await db.users.find_one({"email": "adneia@weconnect360.com.br"}, {"_id": 0})
                    if user:
                        notif = {
                            "id": str(uuid.uuid4()),
                            "tipo": "import_concluida",
                            "titulo": "Importação de Estoque Concluída",
                            "mensagem": f"A importação iniciada por {user_name} foi concluída. {forn_count} fornecedores, {est_inserted} itens novos, {est_updated} atualizados.",
                            "destinatario_email": user['email'],
                            "dados_extras": {"import_id": import_id, "inserted": est_inserted, "updated": est_updated, "forn_count": forn_count},
                            "data_criacao": datetime.now(timezone.utc).isoformat(),
                            "lida": False,
                            "criado_por_nome": "Sistema"
                        }
                        await db.notifications.insert_one(notif)
                except Exception as e:
                    logger.error(f"Erro ao criar notificação: {e}")
                return

            # Arquivo normal — ler aba Tabelão ou primeira aba
            if 'Tabelão' in sheet_names:
                df = pd.read_excel(excel_file, sheet_name='Tabelão')
            else:
                df = pd.read_excel(excel_file, sheet_name=0)

            # Importar Fornecedores se existir junto com Tabelão
            if 'Fornecedores' in sheet_names:
                df_forn = pd.read_excel(excel_file, sheet_name='Fornecedores')
                df_forn.columns = df_forn.columns.str.strip().str.lower()
                for _, row in df_forn.iterrows():
                    fornecedor = None
                    for col in ['fornecedor', 'nome', 'nome_fornecedor']:
                        if col in df_forn.columns:
                            fornecedor = str(row.get(col, '')).strip(); break
                    dias_extras = 5
                    for col in ['dias extras padrão (dias úteis)', 'dias extras padrão', 'dias extras', 'dias_extras']:
                        if col in df_forn.columns:
                            val = row.get(col)
                            if pd.notna(val):
                                try: dias_extras = int(val)
                                except: pass
                            break
                    if fornecedor and fornecedor.lower() not in ('nan', ''):
                        await db.fornecedores.update_one(
                            {"nome": fornecedor},
                            {"$set": {"nome": fornecedor, "dias_extras_padrao": dias_extras, "ultima_atualizacao": datetime.now(timezone.utc).isoformat()}},
                            upsert=True
                        )

        total = len(df)
        await db.import_status.update_one(
            {"import_id": import_id},
            {"$set": {"total": total, "total_rows": total}}
        )

        column_mapping = get_column_mapping()
        original_columns = {col.lower().strip(): col for col in df.columns}
        df.columns = [col.lower().strip() for col in df.columns]

        data_limite = datetime(2025, 1, 1, tzinfo=timezone.utc)
        inserted = updated = skipped = errors = 0

        for idx, row in df.iterrows():
            try:
                pedido_data = extract_pedido_data(row, column_mapping, original_columns)
                numero_pedido = pedido_data.get('numero_pedido', '')
                if not numero_pedido or numero_pedido == 'nan' or numero_pedido == '-':
                    skipped += 1
                    continue
                if should_skip_old_pedido(pedido_data, data_limite):
                    skipped += 1
                    continue
                existing = await db.pedidos_erp.find_one({"numero_pedido": numero_pedido})
                if existing:
                    await db.pedidos_erp.update_one(
                        {"numero_pedido": numero_pedido},
                        {"$set": pedido_data}
                    )
                    updated += 1
                else:
                    pedido_data['id'] = str(uuid.uuid4())
                    pedido_data['imported_at'] = datetime.now(timezone.utc).isoformat()
                    pedido_data['imported_by'] = user_name
                    await db.pedidos_erp.insert_one(pedido_data)
                    inserted += 1

                if (idx + 1) % 100 == 0:
                    progress = int(((idx + 1) / total) * 100)
                    await db.import_status.update_one(
                        {"import_id": import_id},
                        {"$set": {"progress": progress, "processed": idx + 1, "inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}}
                    )
            except Exception as e:
                errors += 1
                logger.error(f"Error importing row {idx}: {e}")

        try:
            await atualizar_motivos_pendencia_automatico()
        except Exception as e:
            logger.error(f"Erro na atualização automática de motivos: {e}")

        await db.import_status.update_one(
            {"import_id": import_id},
            {"$set": {
                "status": "completed",
                "progress": 100,
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }}
        )

        # Notificar admin e equipe de atendimento
        try:
            emails_notificar = [
                "adneia@weconnect360.com.br",
                "caio@weconnect360.com.br",
                "leticia@weconnect360.com.br",
            ]
            usuarios = await db.users.find({"email": {"$in": emails_notificar}}, {"_id": 0}).to_list(10)
            for usuario in usuarios:
                notif = {
                    "id": str(uuid.uuid4()),
                    "tipo": "import_concluida",
                    "titulo": "Tabelão Atualizado",
                    "mensagem": f"A base de pedidos foi atualizada por {user_name}. {inserted} novos, {updated} atualizados, {skipped} ignorados, {errors} erros.",
                    "destinatario_email": usuario['email'],
                    "dados_extras": {"import_id": import_id, "inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors},
                    "data_criacao": datetime.now(timezone.utc).isoformat(),
                    "lida": False,
                    "criado_por_nome": "Sistema"
                }
                await db.notifications.insert_one(notif)
        except Exception as e:
            logger.error(f"Erro ao criar notificação: {e}")

        logger.info(f"Import completed: {inserted} inserted, {updated} updated, {skipped} skipped, {errors} errors")

    except Exception as e:
        logger.error(f"Import error: {e}")
        import traceback
        logger.error(traceback.format_exc())


@router.post("/pedidos-erp/import", response_model=dict)
async def import_pedidos(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Formato de arquivo inválido. Use .xlsx, .xls ou .csv")
    content = await file.read()

    # Pré-calcular total de linhas e gerar import_id
    import_id = str(uuid.uuid4())[:8]
    total_rows = 0
    try:
        from io import BytesIO
        import pandas as pd
        if file.filename.endswith('.csv'):
            # Para CSV, contar linhas rápido sem ler todo o conteúdo
            total_rows = content.count(b'\n')
        else:
            # Para Excel, ler apenas header para estimar (openpyxl read_only)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                total_rows = ws.max_row - 1 if ws.max_row else 0  # -1 para header
                wb.close()
            except Exception:
                total_rows = 0  # Se falhar, será calculado na task
    except Exception as e:
        logger.error(f"Erro ao pré-calcular linhas: {e}")
        total_rows = 0

    background_tasks.add_task(process_import_background, content, file.filename, current_user['name'], current_user['email'], import_id)
    return {
        "message": "Importação iniciada em background",
        "status": "processing",
        "import_id": import_id,
        "total_rows": total_rows
    }


# ============== ESTOQUE ==============

@router.get("/estoque")
async def get_estoque(
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"id_item": search_regex},
            {"descricao": search_regex},
            {"nome": search_regex}
        ]
    skip = (page - 1) * page_size
    total = await db.estoque_sigeq.count_documents(query)
    items = await db.estoque_sigeq.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/estoque/{item_id}")
async def get_estoque_item(item_id: str, current_user: dict = Depends(get_current_user)):
    item = await db.estoque_sigeq.find_one({"id_item": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado no estoque")
    return item


# ============== FORNECEDORES ==============

@router.get("/fornecedores")
async def list_fornecedores(current_user: dict = Depends(get_current_user)):
    return await db.fornecedores.find({}, {"_id": 0}).sort("nome", 1).to_list(100)


@router.post("/fornecedores")
async def create_fornecedor(data: dict, current_user: dict = Depends(get_current_user)):
    data['id'] = str(uuid.uuid4())
    data['criado_por'] = current_user['name']
    data['criado_em'] = datetime.now(timezone.utc).isoformat()
    await db.fornecedores.insert_one(data)
    return {"message": "Fornecedor criado com sucesso", "id": data['id']}


@router.put("/fornecedores/{fornecedor_id}")
async def update_fornecedor(fornecedor_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    result = await db.fornecedores.update_one({"id": fornecedor_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
    return {"message": "Fornecedor atualizado com sucesso"}


# ============== IMPORT ESTOQUE/FORNECEDORES SHEETS ==============

@router.post("/admin/sync-fornecedores")
async def sync_fornecedores(current_user: dict = Depends(get_current_user)):
    try:
        from google_sheets import sheets_client
        dados = sheets_client.get_fornecedores_data()
        if not dados:
            return {"success": False, "message": "Nenhum dado encontrado na planilha"}
        inserted = updated = 0
        for item in dados:
            nome = item.get('nome', '').strip()
            if not nome:
                continue
            existing = await db.fornecedores.find_one({"nome": nome})
            if existing:
                await db.fornecedores.update_one({"nome": nome}, {"$set": item})
                updated += 1
            else:
                item['id'] = str(uuid.uuid4())
                await db.fornecedores.insert_one(item)
                inserted += 1
        return {"success": True, "message": f"Fornecedores sincronizados: {inserted} novos, {updated} atualizados"}
    except Exception as e:
        logger.error(f"Erro ao sincronizar fornecedores: {e}")
        return {"success": False, "message": str(e)}


@router.post("/admin/sync-transportadoras-devolucoes")
async def sync_transportadoras_devolucoes(current_user: dict = Depends(get_current_user)):
    try:
        from google_sheets import sheets_client
        dados = sheets_client.get_transportadoras_devolucoes_data()
        if not dados:
            return {"success": False, "message": "Nenhum dado encontrado"}
        inserted = 0
        for item in dados:
            existing = await db.transportadoras_devolucoes.find_one({"codigo": item.get('codigo')})
            if not existing:
                item['id'] = str(uuid.uuid4())
                await db.transportadoras_devolucoes.insert_one(item)
                inserted += 1
        return {"success": True, "message": f"{inserted} transportadoras importadas"}
    except Exception as e:
        logger.error(f"Erro: {e}")
        return {"success": False, "message": str(e)}
