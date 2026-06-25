"""
Módulo: Base J&T
Importação e consulta do código JMS da J&T Express.

Arquivo Excel (.xlsx) com colunas:
- "Encomenda No." → número do pedido (chave)
- "Número de pedido JMS" → código JMS
- "Status" → status da entrega
- "Previsão de Entrega" → data prevista

Lógica de importação:
- JMS igual ao anterior → só atualiza status/previsão
- JMS diferente do anterior → guarda jms_anterior, usa novo JMS
- Pedido não presente no arquivo → marca descricao_status = "JMS EXCLUÍDO"

Observação: o código de rastreio (chave de acesso) NÃO é alterado por este import.
JMS é usado apenas para abertura de tickets internos.
"""
import io
import uuid
import logging
from fastapi import APIRouter, Depends, UploadFile, File
from datetime import datetime, timezone
from utils.auth import get_current_user
from utils.database import db
from pymongo import UpdateOne

router = APIRouter()
logger = logging.getLogger(__name__)


def _norm(s):
    """Normaliza string: trim, lower, sem acentos básicos."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # remove acentos comuns para casar headers
    for old, new in [("á", "a"), ("ã", "a"), ("â", "a"), ("à", "a"),
                     ("é", "e"), ("ê", "e"), ("í", "i"), ("ó", "o"),
                     ("ô", "o"), ("õ", "o"), ("ú", "u"), ("ç", "c")]:
        s = s.replace(old, new)
    return s


def _find_col_idx(headers, *candidatos):
    """Encontra o índice da coluna pelo nome (case-insensitive, sem acentos)."""
    headers_norm = [_norm(h) for h in headers]
    for cand in candidatos:
        cand_n = _norm(cand)
        for i, h in enumerate(headers_norm):
            if cand_n in h:
                return i
    return -1


@router.post("/api/base-jt/importar")
async def importar_base_jt(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Importa o Excel da Base J&T (xlsx).
    - JMS novo → atualiza, guarda anterior em jms_anterior
    - Pedido ausente no arquivo → marca como JMS EXCLUÍDO
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"ok": False, "message": "openpyxl não instalado"}

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return {"ok": False, "message": f"Erro ao abrir Excel: {e}"}

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return {"ok": False, "message": "Arquivo vazio"}

    # Mapear colunas pelos nomes (mais robusto que índice fixo)
    idx_pedido = _find_col_idx(headers, "Encomenda No", "Numero pedido", "Pedido")
    idx_jms = _find_col_idx(headers, "Numero de pedido JMS", "JMS", "pedido JMS")
    idx_status = _find_col_idx(headers, "Status")
    idx_previsao = _find_col_idx(headers, "Previsao de Entrega", "Previsao Entrega")

    if idx_pedido < 0 or idx_jms < 0:
        return {
            "ok": False,
            "message": f"Colunas obrigatórias não encontradas (Encomenda No / JMS). Headers: {headers[:5]}...",
        }

    # 1. Carregar estado atual da base
    existentes = await db.base_jt.find(
        {}, {"numero_pedido": 1, "jms": 1, "_id": 0}
    ).to_list(None)
    jms_atual = {d["numero_pedido"]: d.get("jms", "") for d in existentes}

    # 2. Processar linhas
    ops = []
    pedidos_no_arquivo = set()
    count = 0
    novos = 0
    atualizados = 0
    atualizados_jms = 0
    erros = 0

    agora = datetime.now(timezone.utc).isoformat()

    for row in rows_iter:
        if not row or idx_pedido >= len(row):
            erros += 1
            continue
        try:
            numero_pedido = str(row[idx_pedido] or "").strip().split(".")[0]
            jms_raw = row[idx_jms] if idx_jms >= 0 and idx_jms < len(row) else None
            jms = str(jms_raw or "").strip().split(".")[0]

            if not numero_pedido or not jms or numero_pedido == "0":
                erros += 1
                continue

            pedidos_no_arquivo.add(numero_pedido)

            status = ""
            if 0 <= idx_status < len(row) and row[idx_status] is not None:
                status = str(row[idx_status]).strip()

            previsao = ""
            if 0 <= idx_previsao < len(row) and row[idx_previsao] is not None:
                val = row[idx_previsao]
                if hasattr(val, "isoformat"):
                    previsao = val.strftime("%Y-%m-%d")
                else:
                    previsao = str(val).strip()

            doc = {
                "numero_pedido": numero_pedido,
                "jms": jms,
                "descricao_status": status,
                "previsao_entrega": previsao,
                "updated_at": agora,
            }

            # Conta novos vs atualizados
            ja_existia = numero_pedido in jms_atual
            if ja_existia:
                atualizados += 1
                # Se JMS mudou, guardar o anterior
                jms_antigo = jms_atual.get(numero_pedido)
                if jms_antigo and jms_antigo != jms:
                    doc["jms_anterior"] = jms_antigo
                    atualizados_jms += 1
            else:
                novos += 1

            ops.append(UpdateOne(
                {"numero_pedido": numero_pedido},
                {
                    "$set": doc,
                    "$setOnInsert": {"criado_em": agora},
                },
                upsert=True,
            ))
            count += 1

            if len(ops) >= 500:
                await db.base_jt.bulk_write(ops, ordered=False)
                ops = []
        except Exception as e:
            logger.warning(f"Erro na linha J&T: {e}")
            erros += 1
            continue

    if ops:
        await db.base_jt.bulk_write(ops, ordered=False)

    # 3. NÃO marca mais como "JMS EXCLUÍDO" automaticamente.
    #    Mesma lógica do base_total: arquivos podem ser parciais/filtrados,
    #    então marcar EXCLUÍDO automático causaria perda de dados válidos.
    excluidos = 0

    # 4. Índice único
    try:
        await db.base_jt.create_index("numero_pedido", unique=True)
    except Exception:
        pass

    # 5. Notificar
    try:
        user_name = current_user.get("name") or current_user.get("email", "Sistema")
        partes = [f"{count} registros processados"]
        if novos:
            partes.append(f"{novos} novos")
        if atualizados:
            partes.append(f"{atualizados} atualizados")
        if atualizados_jms:
            partes.append(f"{atualizados_jms} JMS reemitido")
        if excluidos:
            partes.append(f"{excluidos} marcados como JMS Excluído")
        if erros:
            partes.append(f"{erros} ignorados")
        mensagem = f"Base J&T importada por {user_name}. {', '.join(partes)}."

        usuarios = await db.users.find({}, {"email": 1, "_id": 0}).to_list(50)
        for usuario in usuarios:
            notif = {
                "id": str(uuid.uuid4()),
                "tipo": "import_concluida",
                "titulo": "Base J&T Atualizada",
                "mensagem": mensagem,
                "destinatario_email": usuario["email"],
                "dados_extras": {
                    "importados": count,
                    "novos": novos,
                    "atualizados": atualizados,
                    "atualizados_jms": atualizados_jms,
                    "excluidos": excluidos,
                    "erros": erros,
                },
                "data_criacao": datetime.now(timezone.utc).isoformat(),
                "lida": False,
                "criado_por_nome": "Sistema",
            }
            await db.notifications.insert_one(notif)
    except Exception as e:
        logger.warning(f"Erro ao criar notificação Base J&T: {e}")

    # Monta mensagem de resposta detalhada
    msg_partes = [f"{count} registros processados"]
    msg_partes.append(f"{novos} novos")
    msg_partes.append(f"{atualizados} atualizados")
    msg_partes.append(f"{atualizados_jms} JMS reemitido")
    if excluidos:
        msg_partes.append(f"{excluidos} excluídos")
    msg_partes.append(f"{erros} ignorados")

    return {
        "ok": True,
        "importados": count,
        "novos": novos,
        "atualizados": atualizados,
        "atualizados_jms": atualizados_jms,
        "excluidos": excluidos,
        "erros": erros,
        "message": f"{msg_partes[0]} ({', '.join(msg_partes[1:])}).",
    }


@router.get("/api/base-jt/{numero_pedido}")
async def get_jms_by_pedido(
    numero_pedido: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Retorna JMS e status da J&T para um número de pedido.
    """
    pedido = str(numero_pedido).strip().split(".")[0]
    doc = await db.base_jt.find_one({"numero_pedido": pedido}, {"_id": 0})
    if not doc:
        return {"numero_pedido": pedido, "jms": None, "descricao_status": None}
    return doc
