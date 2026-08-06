"""
Módulo: Cancelamentos
Gerencia cancelamentos de pedidos com 3 fluxos:
- AES (Compras): solicitação por falta de estoque no fornecedor
- ETR (Produção): identificado na produção (perda, quebra, falha cadastro)
- Erro na Nota: nota rejeitada, requer acionar cliente + lançar nova entrega
"""
import uuid
import logging
import re
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from datetime import datetime, timezone, timedelta
from utils.auth import get_current_user
from utils.database import db

router = APIRouter(prefix="/api")
logger = logging.getLogger(__name__)

# Cache do breakdown de lifecycle dos cards (recalcular é caro: lê ~1.6k docs + bigdata)
_lifecycle_cache = {"ts": None, "data": None}
_LIFECYCLE_TTL = 120  # segundos


# ============== MODELS ==============
class CancelamentoCreate(BaseModel):
    tipo: str = Field(..., description="aes | etr | erro_nota")
    numero_pedido: str  # Entrega
    motivo: Optional[str] = None
    acao: Optional[str] = None  # Cancelar / Reenviar
    motivo_rejeicao: Optional[str] = None  # para erro_nota
    ticket: Optional[str] = None
    instancia: Optional[str] = None
    zerado_reserva: Optional[bool] = None
    observacao: Optional[str] = None
    # Para erro_nota
    cliente_confirmado_nome: Optional[str] = None
    cliente_confirmado_cpf: Optional[str] = None
    cliente_confirmado_endereco: Optional[str] = None
    nova_entrega: Optional[str] = None


class CancelamentoUpdate(BaseModel):
    motivo: Optional[str] = None
    acao: Optional[str] = None
    motivo_rejeicao: Optional[str] = None
    ticket: Optional[str] = None
    instancia: Optional[str] = None
    zerado_reserva: Optional[bool] = None
    observacao: Optional[str] = None
    cliente_confirmado_nome: Optional[str] = None
    cliente_confirmado_cpf: Optional[str] = None
    cliente_confirmado_endereco: Optional[str] = None
    nova_entrega: Optional[str] = None
    data_encerramento: Optional[str] = None
    status: Optional[str] = None  # pendente | em_andamento | encerrado
    prioridade: Optional[bool] = None
    sku_similar: Optional[str] = None   # SKU do item similar (AES)
    nome_similar: Optional[str] = None  # Nome do produto similar (auto-preenchido)
    id_similar: Optional[str] = None    # ID BSeller do produto similar (auto-preenchido)
    em_compras: Optional[bool] = None   # True = item movido para a caixa Compras (acionando estoque)
    analise_similar: Optional[str] = None  # pendente | proposto | cancelar | sem_similar
    similares_propostos: Optional[list] = None  # lista de SKUs escolhidos pelo analista


def _br_now():
    return datetime.now(timezone(timedelta(hours=-3)))


def _iso_now_utc():
    return datetime.now(timezone.utc).isoformat()


# Statuses do pedido (em pedidos_erp.status_pedido) que MANTÊM um cancelamento AES como 'pendente'.
# Regra de negócio: cancelamento AES é criado quando há falta de estoque (status "Aguardando estoque").
# Ao acionar o cancelamento, o pedido é cancelado no parceiro → status vai pra "Cancelado/Cancelada".
# Qualquer outro status indica que o pedido avançou na esteira (NF emitida, enviado, entregue, etc.)
# e o cancelamento não pode mais ser efetivado → deve ser encerrado automaticamente.
# Threshold para tratar XD como "consta estoque" — gera alerta + nota na observação + entra na relação Compras.
_XD_THRESHOLD = 10

# Statuses do pedido que indicam que cancelamento já foi consumado (não adianta acionar Compras)
_AES_STATUS_JA_CANCELADO = {"Cancelado", "Cancelada", "CANCELADO", "Entrega cancelada"}


_AES_STATUS_MANTER_PENDENTE = {
    "",  # sem dados de status — não temos info pra decidir
    "Aguardando estoque",
    # Acionamento de cancelamento foi bem-sucedido — usuário faz a baixa manualmente
    # através do encerramento do atendimento (chamado). Não auto-fechar pelo status do pedido.
    "Cancelado", "Cancelada", "CANCELADO", "Entrega cancelada",
    # Pré-faturamento — ainda no início do fluxo, cancelamento viável
    "Pedido Incluido", "Pagamento OK", "Pedido aprovado",
    "Aguardando aprovação de pagamento", "Aguardando liberacao do SAC",
}


async def _auto_encerrar_aes_por_status_pedido() -> dict:
    """
    Encerra automaticamente cancelamentos AES pendentes em DUAS situações:

    1) Pedido avançou para status que invalida o cancelamento (NFe Aprovada,
       Entregue a Transportadora, EM TRÂNSITO, Entregue ao Cliente, etc.) →
       cancelamento falhou, o pedido foi adiante. Encerra com nota indicando
       o status novo.
    2) Chamado/atendimento vinculado (mesmo numero_pedido) foi encerrado
       (pendente=false). Sinal de que o time já tratou e finalizou o caso.
       Encerra com nota indicando o id do atendimento encerrado.

    Status do pedido que NÃO disparam encerramento (whitelist):
        Aguardando estoque, Cancelado/Cancelada/CANCELADO/Entrega cancelada,
        pré-faturamento (Pedido Incluido, Pagamento OK, etc.).
    Idempotente: cancelamentos já encerrados não são re-processados.
    """
    pipeline = [
        {"$match": {"tipo": "aes", "status": "pendente"}},
        # Lookup status atual no tabelão
        {"$lookup": {
            "from": "pedidos_erp",
            "localField": "numero_pedido",
            "foreignField": "numero_pedido",
            "as": "_pe",
        }},
        # Lookup chamado/atendimento vinculado
        {"$lookup": {
            "from": "chamados",
            "localField": "numero_pedido",
            "foreignField": "numero_pedido",
            "as": "_chamado",
        }},
        {"$addFields": {
            "_status_atual": {"$ifNull": [{"$arrayElemAt": ["$_pe.status_pedido", 0]}, ""]},
            "_chamado_pendente": {"$arrayElemAt": ["$_chamado.pendente", 0]},
            "_chamado_id_atd": {"$arrayElemAt": ["$_chamado.id_atendimento", 0]},
            "_tem_ticket": {"$ne": [{"$ifNull": ["$ticket", ""]}, ""]},
        }},
        # Filtra os elegíveis: pedido avançou OU chamado encerrado
        {"$match": {
            "$or": [
                {"_status_atual": {"$nin": list(_AES_STATUS_MANTER_PENDENTE)}},
                {"_chamado_pendente": False},
            ]
        }},
        {"$project": {
            "_id": 0, "id": 1, "numero_pedido": 1, "observacao": 1,
            "_status_atual": 1, "_chamado_pendente": 1, "_chamado_id_atd": 1, "_tem_ticket": 1,
        }},
    ]
    docs = await db.cancelamentos.aggregate(pipeline).to_list(1000)
    if not docs:
        return {"encerrados": 0, "ids": [], "alertados": 0}

    now_iso = _iso_now_utc()
    data_br_curta = _br_now().strftime("%d/%m")
    data_br_completa = _br_now().strftime("%Y-%m-%d")
    ids_encerrados = []
    alertados = 0
    for d in docs:
        novo_status = d.get("_status_atual", "") or ""
        chamado_encerrado = d.get("_chamado_pendente") is False
        id_atd = d.get("_chamado_id_atd") or ""
        tem_ticket = d.get("_tem_ticket") is True
        obs_atual = d.get("observacao", "") or ""
        status_eh_invalido = novo_status not in _AES_STATUS_MANTER_PENDENTE

        # REGRA ETR: pedido movimentou para 'Entregue a Transportadora'. Mesmo com
        # ticket aberto, encerra o cancelamento (o pedido seguiu) e marca para o card
        # avisar o atendente a conferir se o cliente foi acionado.
        if novo_status.strip().lower() == "entregue a transportadora":
            nota = f"{data_br_curta} - encerrado devido a movimentação para ETR"
            nova_obs = (nota + ("\n" + obs_atual if obs_atual else "")).strip() if nota not in obs_atual else obs_atual
            await db.cancelamentos.update_one({"id": d["id"]}, {"$set": {
                "status": "encerrado",
                "data_encerramento": data_br_completa,
                "observacao": nova_obs,
                "encerrado_automaticamente": True,
                "encerrado_por_etr": True,
                "status_pedido": novo_status,
                "updated_at": now_iso,
            }})
            ids_encerrados.append(d["id"])
            continue

        # CASO ESPECIAL: pedido MOVIMENTOU mas há TICKET de cancelamento aberto.
        # Não encerra — mantém pendente e alerta, pois há acionamento ativo a acompanhar.
        # (só vale quando o gatilho foi o status do pedido, não o chamado encerrado)
        if status_eh_invalido and tem_ticket and not chamado_encerrado:
            nota = f"{data_br_curta} - ⚠️ Pedido movimentou para '{novo_status}' com ticket aberto — verificar acionamento com o canal/transportadora"
            if "movimentou" not in obs_atual or nota not in obs_atual:
                if nota not in obs_atual:
                    nova_obs = (nota + ("\n" + obs_atual if obs_atual else "")).strip()
                    await db.cancelamentos.update_one(
                        {"id": d["id"]},
                        {"$set": {"observacao": nova_obs, "movimentou_com_ticket": True, "updated_at": now_iso}},
                    )
                    alertados += 1
            continue

        # Define a nota conforme o gatilho. Chamado encerrado tem prioridade
        # se o status do pedido ainda for válido (whitelist).
        if chamado_encerrado and not status_eh_invalido:
            nota = f"{data_br_curta} - Encerrado automaticamente: atendimento {id_atd} foi encerrado"
        else:
            nota = f"{data_br_curta} - Encerrado automaticamente: pedido alterado para '{novo_status}'"

        if nota not in obs_atual:
            nova_obs = (nota + ("\n" + obs_atual if obs_atual else "")).strip()
        else:
            nova_obs = obs_atual

        update_doc = {
            "status": "encerrado",
            "data_encerramento": data_br_completa,
            "observacao": nova_obs,
            "encerrado_automaticamente": True,
            "updated_at": now_iso,
        }
        if novo_status:
            update_doc["status_pedido"] = novo_status

        await db.cancelamentos.update_one({"id": d["id"]}, {"$set": update_doc})
        ids_encerrados.append(d["id"])
    logger.info(f"[auto-encerrar AES] {len(ids_encerrados)} encerrado(s), {alertados} mantido(s) com alerta (ticket aberto + movimentou)")
    return {"encerrados": len(ids_encerrados), "ids": ids_encerrados, "alertados": alertados}


async def _auto_nota_consta_estoque() -> dict:
    """
    Para AES pendentes que TÊM ticket aberto, estoque XD > _XD_THRESHOLD
    e pedido AINDA não cancelado, adiciona nota na observação:
    "DD/MM - Consta estoque, verificar possibilidade de segurar o cancelamento."
    Idempotente: não duplica.
    """
    pipeline = [
        {"$match": {"tipo": "aes", "status": "pendente", "em_compras": {"$ne": True}}},
        {"$lookup": {
            "from": "estoque_sigeq",
            "let": {"sku": "$codigo_item_vtex"},
            "pipeline": [
                {"$match": {"$expr": {"$and": [
                    {"$eq": ["$cod_terceiro", "$$sku"]},
                    {"$eq": ["$source", "SIGEQ425"]},
                ]}}},
                {"$limit": 1},
                {"$project": {"_id": 0, "disp_venda": 1}},
            ],
            "as": "_xd",
        }},
        {"$addFields": {
            "_xd_disp": {"$ifNull": [{"$arrayElemAt": ["$_xd.disp_venda", 0]}, 0]},
            "_ticket_preenchido": {"$ne": [{"$ifNull": ["$ticket", ""]}, ""]},
        }},
        {"$match": {
            "_xd_disp": {"$gt": _XD_THRESHOLD},
            "_ticket_preenchido": True,
        }},
        {"$project": {"_id": 0, "id": 1, "observacao": 1}},
    ]
    docs = await db.cancelamentos.aggregate(pipeline).to_list(2000)
    if not docs:
        return {"adicionadas": 0}
    data_br = _br_now().strftime("%d/%m")
    nota_texto = f"{data_br} - Consta estoque, verificar possibilidade de segurar o cancelamento."
    adicionadas = 0
    for d in docs:
        obs = d.get("observacao") or ""
        if "Consta estoque, verificar possibilidade de segurar" in obs:
            continue
        nova_obs = (nota_texto + ("\n" + obs if obs else "")).strip()
        await db.cancelamentos.update_one(
            {"id": d["id"]},
            {"$set": {"observacao": nova_obs, "updated_at": _iso_now_utc()}},
        )
        adicionadas += 1
    if adicionadas:
        logger.info(f"[auto-nota consta-estoque] {adicionadas} nota(s) adicionada(s)")
    return {"adicionadas": adicionadas}


async def _analisar_similares_pendentes(limit: int = 200, ignorar_frescor: bool = False) -> dict:
    """
    Para AES pendentes ainda NÃO analisados (sem campo analise_similar ou em estado
    recomputável), busca similares no catálogo (mesma tensão, com estoque).
    - Encontrou similares → analise_similar='pendente', grava similares_sugeridos.
    - Não encontrou → analise_similar='sem_similar' (segue como cancelamento normal).
    Não mexe em quem o analista já decidiu (proposto / cancelar).
    Idempotente. Pesado (Postgres) — chamar em lote controlado (criação, backlog, sync).
    Frescor: pula quem já foi analisado há menos de 1h (o sync de 15min reprocessava os
    mesmos ~190 docs toda rodada). Doc novo (sem o campo) é analisado no próximo sync.
    ignorar_frescor=True força rodada completa (endpoint manual).
    """
    from routes.produtos_busca import computar_similares_compactos
    # Estados que devem ser (re)analisados: ainda não decididos pelo analista.
    # Pula quem já é Similar (analista escolheu manualmente) — esses não viram picker.
    query = {
        "tipo": "aes",
        "status": "pendente",
        "acao": {"$not": {"$regex": "similar", "$options": "i"}},
        "$or": [
            {"analise_similar": {"$exists": False}},
            {"analise_similar": {"$in": ["pendente", "sem_similar"]}},
        ],
    }
    if not ignorar_frescor:
        cutoff_frescor = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
        query["$and"] = [{"$or": [
            {"similares_analisado_em": {"$exists": False}},
            {"similares_analisado_em": {"$lt": cutoff_frescor}},
        ]}]
    docs = await db.cancelamentos.find(
        query, {"_id": 0, "id": 1, "codigo_item_vtex": 1, "numero_pedido": 1}
    ).to_list(limit)
    if not docs:
        return {"analisados": 0, "com_similar": 0}

    analisados = 0
    com_similar = 0
    for d in docs:
        sku = d.get("codigo_item_vtex") or ""
        if not sku:
            await db.cancelamentos.update_one(
                {"id": d["id"]},
                {"$set": {"analise_similar": "sem_similar", "similares_sugeridos": [], "similares_analisado_em": _iso_now_utc(), "updated_at": _iso_now_utc()}},
            )
            analisados += 1
            continue
        res = await computar_similares_compactos(sku, d.get("numero_pedido"), max_propostos=5)
        if res.get("found"):
            await db.cancelamentos.update_one(
                {"id": d["id"]},
                {"$set": {
                    "analise_similar": "pendente",
                    "similares_sugeridos": res["propostos"],
                    "similares_analisado_em": _iso_now_utc(),
                    # Congela a posição na faixa "Similar": mesmo que depois o analista
                    # decida cancelar, o item NÃO cai pro fim (não some de vista).
                    "entrou_similar": True,
                    "updated_at": _iso_now_utc(),
                }},
            )
            com_similar += 1
        else:
            await db.cancelamentos.update_one(
                {"id": d["id"]},
                {"$set": {"analise_similar": "sem_similar", "similares_sugeridos": [], "similares_analisado_em": _iso_now_utc(), "updated_at": _iso_now_utc()}},
            )
        analisados += 1
    logger.info(f"[analisar-similares] {analisados} analisados, {com_similar} com similar")
    return {"analisados": analisados, "com_similar": com_similar}


async def _auto_voltar_de_compras() -> dict:
    """
    Para cancelamentos AES marcados como em_compras=true, desfaz a marcação se:
      - XD zerou (não tem mais estoque pra acionar), OU
      - Ticket foi preenchido (atendente acionou o canal — agora aguarda retorno).
    Adiciona nota na observação explicando a razão.
    """
    pipeline = [
        {"$match": {"tipo": "aes", "status": "pendente", "em_compras": True}},
        {"$lookup": {
            "from": "estoque_sigeq",
            "let": {"sku": "$codigo_item_vtex"},
            "pipeline": [
                {"$match": {"$expr": {"$and": [
                    {"$eq": ["$cod_terceiro", "$$sku"]},
                    {"$eq": ["$source", "SIGEQ425"]},
                ]}}},
                {"$limit": 1},
                {"$project": {"_id": 0, "disp_venda": 1}},
            ],
            "as": "_xd",
        }},
        {"$addFields": {
            "_xd_disp": {"$ifNull": [{"$arrayElemAt": ["$_xd.disp_venda", 0]}, 0]},
            "_ticket_str": {"$ifNull": ["$ticket", ""]},
        }},
        {"$match": {"$or": [
            {"_xd_disp": {"$lte": 0}},
            {"_ticket_str": {"$nin": ["", None]}},
        ]}},
        {"$project": {"_id": 0, "id": 1, "observacao": 1, "_xd_disp": 1, "_ticket_str": 1}},
    ]
    docs = await db.cancelamentos.aggregate(pipeline).to_list(2000)
    if not docs:
        return {"voltados": 0}
    data_br = _br_now().strftime("%d/%m")
    voltados = 0
    for d in docs:
        xd = d.get("_xd_disp", 0)
        ticket = (d.get("_ticket_str") or "").strip()
        if ticket:
            nota = f"{data_br} - Compras: ticket aberto ({ticket}), aguardando retorno do canal"
        else:
            nota = f"{data_br} - Compras: estoque zerou, voltou para o canal"
        obs = d.get("observacao") or ""
        nova_obs = obs if nota in obs else (nota + ("\n" + obs if obs else "")).strip()
        await db.cancelamentos.update_one(
            {"id": d["id"]},
            {"$set": {"observacao": nova_obs, "em_compras": False, "updated_at": _iso_now_utc()}},
        )
        voltados += 1
    if voltados:
        logger.info(f"[auto-voltar Compras] {voltados} cancelamento(s) retornaram ao canal")
    return {"voltados": voltados}


async def _auto_mover_para_compras() -> dict:
    """
    Para cada AES pendente cujo SKU tem estoque cross-dock > _XD_THRESHOLD
    e cujo TICKET ainda está vazio (canal não foi acionado):
      - Marca em_compras=true (move para a caixa Compras)
      - Adiciona nota: "DD/MM - Acionar compras para verificar o estoque"
    Regra olha SOMENTE o ticket (status do pedido não importa — se houver
    estoque pode-se subir pedido manual mesmo que o pedido esteja cancelado).
    Idempotente: só age em quem ainda não está em_compras.
    """
    pipeline = [
        {"$match": {"tipo": "aes", "status": "pendente", "em_compras": {"$ne": True}}},
        # XD lookup
        {"$lookup": {
            "from": "estoque_sigeq",
            "let": {"sku": "$codigo_item_vtex"},
            "pipeline": [
                {"$match": {"$expr": {"$and": [
                    {"$eq": ["$cod_terceiro", "$$sku"]},
                    {"$eq": ["$source", "SIGEQ425"]},
                ]}}},
                {"$limit": 1},
                {"$project": {"_id": 0, "disp_venda": 1}},
            ],
            "as": "_xd",
        }},
        {"$addFields": {
            "_xd_disp": {"$ifNull": [{"$arrayElemAt": ["$_xd.disp_venda", 0]}, 0]},
            # Ticket considerado "vazio" se for null, ausente ou string em branco
            "_ticket_vazio": {"$or": [
                {"$eq": [{"$ifNull": ["$ticket", ""]}, ""]},
                {"$eq": [{"$type": "$ticket"}, "missing"]},
            ]},
        }},
        {"$match": {
            "_xd_disp": {"$gt": _XD_THRESHOLD},
            "_ticket_vazio": True,
        }},
        {"$project": {"_id": 0, "id": 1, "observacao": 1, "_xd_disp": 1}},
    ]
    docs = await db.cancelamentos.aggregate(pipeline).to_list(2000)
    if not docs:
        return {"movidos": 0}

    data_br_curta = _br_now().strftime("%d/%m")
    nota_texto = f"{data_br_curta} - Acionar compras para verificar o estoque"
    movidos = 0
    for d in docs:
        obs = d.get("observacao") or ""
        nova_obs = obs if "Acionar compras para verificar o estoque" in obs \
                       else (nota_texto + ("\n" + obs if obs else "")).strip()
        await db.cancelamentos.update_one(
            {"id": d["id"]},
            {"$set": {
                "observacao": nova_obs,
                "em_compras": True,
                "updated_at": _iso_now_utc(),
            }},
        )
        movidos += 1
    if movidos:
        logger.info(f"[auto-mover Compras] {movidos} cancelamento(s) movido(s) automaticamente")
    return {"movidos": movidos}


async def _auto_reabrir_encerrados() -> dict:
    """Reabre AES encerrados quando chega NOVA solicitação (Smart Compras) DEPOIS da
    data de encerramento. Não reabre pedidos realmente cancelados (entrega cancelada).
    Idempotente: ao reabrir vira 'pendente', então não é reprocessado."""
    encerrados = await db.cancelamentos.find(
        {"tipo": "aes", "status": "encerrado", "data_encerramento": {"$nin": ["", None]}},
        {"_id": 0, "id": 1, "numero_pedido": 1, "codigo_item_vtex": 1,
         "data_encerramento": 1, "observacao": 1},
    ).to_list(5000)
    if not encerrados:
        return {"reabertos": 0}
    reabertos = 0
    now_iso = _iso_now_utc()
    data_br = _br_now().strftime("%d/%m")

    # Bulk: avisos ativos do Smart Compras UMA vez (evita 1 find_one por encerrado —
    # com ~900 encerrados isso custava ~2s em toda listagem da tela).
    pedidos_enc = list({(c.get("numero_pedido") or "").strip()
                        for c in encerrados if c.get("numero_pedido")})
    avisos_raw = await db.avisos_compras.find(
        {"numero_pedido": {"$in": pedidos_enc}, "status": {"$in": ["aberto", "faturado"]}},
        {"_id": 0, "numero_pedido": 1, "sku": 1, "atualizado_em": 1},
    ).to_list(20000) if pedidos_enc else []
    avisos_por_ped = {}
    for a in avisos_raw:
        avisos_por_ped.setdefault((a.get("numero_pedido") or "").strip(), []).append(a)

    # Candidatos (aviso novo após o encerramento) — checa o status do pedido em bulk
    candidatos = []
    for c in encerrados:
        ped = (c.get("numero_pedido") or "").strip()
        enc = str(c.get("data_encerramento") or "")[:10]  # YYYY-MM-DD
        if not ped or len(enc) < 10:
            continue
        sku = (c.get("codigo_item_vtex") or "").strip()
        corte = enc + "T23:59:59.999999"
        tem_novo = any(
            (str(a.get("atualizado_em") or "") > corte) and (not sku or (a.get("sku") or "").strip() == sku)
            for a in avisos_por_ped.get(ped, [])
        )
        if tem_novo:
            candidatos.append(c)
    if not candidatos:
        return {"reabertos": 0}
    peds_cand = list({(c.get("numero_pedido") or "").strip() for c in candidatos})
    status_ped = {}
    async for p in db.pedidos_erp.find(
            {"numero_pedido": {"$in": peds_cand}},
            {"_id": 0, "numero_pedido": 1, "status_pedido": 1}):
        status_ped[p["numero_pedido"]] = p.get("status_pedido") or ""

    for c in candidatos:
        ped = (c.get("numero_pedido") or "").strip()
        enc = str(c.get("data_encerramento") or "")[:10]
        # não reabrir se o pedido já foi realmente cancelado (entrega cancelada)
        if status_ped.get(ped, "") in _AES_STATUS_JA_CANCELADO:
            continue
        obs = c.get("observacao", "") or ""
        enc_br = f"{enc[8:10]}/{enc[5:7]}"
        nota = f"{data_br} - Reaberto: nova solicitação no Smart Compras (após encerramento em {enc_br})"
        nova_obs = (nota + ("\n" + obs if obs else "")).strip()
        await db.cancelamentos.update_one(
            {"id": c["id"]},
            {"$set": {"status": "pendente", "data_encerramento": "", "reaberto": True,
                      "reaberto_em": now_iso, "reaberto_motivo": "smart_compras",
                      "observacao": nova_obs, "updated_at": now_iso},
             "$unset": {"encerrado_por_compras": "", "encerrado_por_etr": ""}},
        )
        reabertos += 1
    if reabertos:
        logger.info(f"[auto-reabrir] {reabertos} AES reaberto(s) por nova solicitação Smart Compras")
    return {"reabertos": reabertos}


async def _enrich_from_tabelao(numero_pedido: str, sku: Optional[str] = None) -> dict:
    """Busca dados do pedido no tabelão para preenchimento automático.

    Quando o pedido tem MAIS DE UM ITEM, o pedidos_erp grava um documento por
    (pedido, item). Sem `sku`, find_one() pegava qualquer item — OK pros dados
    do pedido/cliente mas trazia codigo_item_vtex/produto possivelmente de OUTRO
    item. Quando o caller souber o SKU (ex.: fluxo Smart Compras), passa aqui
    pra puxar o item certo.
    """
    pedido = None
    if sku:
        pedido = await db.pedidos_erp.find_one(
            {"numero_pedido": numero_pedido, "codigo_item_vtex": sku},
            {"_id": 0},
        )
    if not pedido:
        pedido = await db.pedidos_erp.find_one({"numero_pedido": numero_pedido}, {"_id": 0})
    if not pedido:
        return {}
    return {
        "canal_vendas": pedido.get("canal_vendas", ""),
        "parceiro": pedido.get("canal_vendas", ""),
        "nome_cliente": pedido.get("nome_cliente", ""),
        "cpf_cliente": pedido.get("cpf_cliente", ""),
        "fone_cliente": pedido.get("fone_cliente", ""),
        "email_cliente": pedido.get("email_cliente", ""),
        "endereco_rua": pedido.get("endereco_rua", ""),
        "endereco_numero": pedido.get("endereco_numero", ""),
        "endereco_complemento": pedido.get("endereco_complemento", ""),
        "endereco_bairro": pedido.get("endereco_bairro", ""),
        "cidade": pedido.get("cidade", ""),
        "uf": pedido.get("uf", ""),
        "cep": pedido.get("cep", ""),
        "produto": pedido.get("produto", ""),
        "codigo_item_bseller": pedido.get("codigo_item_bseller", ""),
        "codigo_item_vtex": pedido.get("codigo_item_vtex", ""),  # SKU/Cód. Terceiro
        "codigo_fornecedor": pedido.get("codigo_fornecedor", ""),
        "departamento": pedido.get("departamento", ""),
        "filial": pedido.get("filial", ""),
        "nota_fiscal": pedido.get("nota_fiscal", ""),
        "serie_nf": pedido.get("serie_nf", ""),
        "chave_nota": pedido.get("chave_nota", ""),
        "preco_final": pedido.get("preco_final", ""),  # Valor da venda
        "quantidade": pedido.get("quantidade", ""),
        "transportadora": pedido.get("transportadora", ""),
        "status_pedido": pedido.get("status_pedido", ""),
        "data_status": pedido.get("data_status", ""),
    }


async def _auto_importar_avisos_smart_compras() -> dict:
    """
    Importa avisos do Smart Compras (db.avisos_compras com status 'aberto'/'faturado'
    e sem cancelamento associado) para o fluxo de Cancelamentos AES.

    Regra (definida pela Adneia):
      - Se o pedido JÁ tem cancelamento AES → mantém o existente e adiciona
        observação "DD/M - cancelamento vindo pelo smart compras" no topo.
      - Se NÃO tem → cria um novo cancelamento AES (status='pendente') com a
        mesma observação + enriquece com dados do pedido_erp.
      - Sem prioridade — segue o fluxo normal.

    Idempotente: marca o aviso com `cancelamento_id` apontando pro cancelamento,
    pra não reprocessar. Avisos já associados a algum cancelamento são pulados.
    """
    stats = {"vistos": 0, "anexados": 0, "criados": 0, "ignorados": 0, "erros": 0}
    try:
        # Avisos ativos ainda não vinculados a cancelamento
        avisos = await db.avisos_compras.find(
            {
                "status": {"$in": ["aberto", "faturado"]},
                "$or": [
                    {"cancelamento_id": {"$exists": False}},
                    {"cancelamento_id": None},
                    {"cancelamento_id": ""},
                ],
            },
            {"_id": 0},
        ).to_list(500)
        stats["vistos"] = len(avisos)
        if not avisos:
            return stats

        data_br_curta = _br_now().strftime("%d/%m")
        nota_smart = f"{data_br_curta} - cancelamento vindo pelo smart compras"

        for av in avisos:
            try:
                numero_pedido = (av.get("numero_pedido") or "").strip()
                sku_aviso = (av.get("sku") or "").strip()
                if not numero_pedido:
                    stats["ignorados"] += 1
                    continue

                # MATCH por (numero_pedido + SKU): um pedido pode ter vários itens com
                # cancelamentos distintos. Casar só por numero_pedido faria o aviso de
                # um item ser anexado ao cancelamento de outro item do mesmo pedido.
                # Casa SÓ pelo codigo_item_vtex (cod_terceiro real, confiável). O
                # codigo_terceiro_planilha vem do parceiro e às vezes traz outro SKU
                # (ex.: cancelamento de Panquequeira com codigo_terceiro_planilha de
                # uma Frigideira que estava na mesma planilha do parceiro).
                filtro_existente = {"numero_pedido": numero_pedido, "tipo": "aes"}
                if sku_aviso:
                    filtro_existente["codigo_item_vtex"] = sku_aviso
                existente = await db.cancelamentos.find_one(
                    filtro_existente,
                    {"_id": 0, "id": 1, "observacao": 1},
                    sort=[("data_criacao", -1)],
                )

                if existente:
                    obs_atual = (existente.get("observacao") or "").strip()
                    if "smart compras" in obs_atual.lower():
                        # já tem a observação — só vincula pra não reprocessar
                        nova_obs = obs_atual
                    elif obs_atual:
                        nova_obs = nota_smart + "\n" + obs_atual
                    else:
                        nova_obs = nota_smart
                    await db.cancelamentos.update_one(
                        {"id": existente["id"]},
                        {"$set": {"observacao": nova_obs, "updated_at": _iso_now_utc()}},
                    )
                    await db.avisos_compras.update_one(
                        {"id": av["id"]},
                        {"$set": {
                            "cancelamento_id": existente["id"],
                            "atualizado_em": _iso_now_utc(),
                        }},
                    )
                    stats["anexados"] += 1
                    logger.info(
                        f"[smart→aes] aviso {av['id']} anexado a cancelamento existente "
                        f"{existente['id']} (pedido {numero_pedido})"
                    )
                else:
                    # Cria novo cancelamento AES, enriquecido com dados do tabelão
                    dados_tabelao = await _enrich_from_tabelao(numero_pedido, sku=sku_aviso)
                    now_br = _br_now()
                    novo_id = str(uuid.uuid4())
                    doc = {
                        "id": novo_id,
                        "tipo": "aes",
                        "status": "pendente",
                        "numero_pedido": numero_pedido,
                        "data_criacao": _iso_now_utc(),
                        "data": now_br.strftime("%Y-%m-%d"),
                        "criado_por": "Smart Compras (auto)",
                        "criado_por_email": av.get("criado_por") or "smartcompras-bot@wct360.com.br",
                        "motivo": av.get("motivo") or "",
                        "acao": "Cancelar",
                        "motivo_rejeicao": "",
                        "ticket": "",
                        "instancia": "",
                        "zerado_reserva": None,
                        "observacao": nota_smart + (("\n" + av["comentario"]) if av.get("comentario") else ""),
                        "cliente_confirmado_nome": "",
                        "cliente_confirmado_cpf": "",
                        "cliente_confirmado_endereco": "",
                        "nova_entrega": "",
                        # Marca a fonte (rastreabilidade)
                        "fonte": "smart_compras",
                        # NOVO: chegou pela migração automática — some da faixa "Novos"
                        # quando o analista decide (propor similar / seguir cancelamento).
                        "novo": True,
                        "aviso_compras_id": av["id"],
                        "aviso_numero_po": av.get("numero_po"),
                        "aviso_po_id": av.get("po_id"),
                        **dados_tabelao,
                        "updated_at": _iso_now_utc(),
                    }
                    # PRIORIDADE: dados do ITEM vêm do AVISO (não do enrich).
                    # O pedidos_erp pode ter só 1 dos N itens do pedido (caso real:
                    # pedido 122023668 só tem a Panquequeira no pedidos_erp, mas
                    # o aviso é sobre a Frigideira BRINL1450). Os dados do PEDIDO
                    # (cliente, endereço, NF, transportadora) do enrich servem,
                    # mas SKU/produto/fornecedor têm que ser sempre os do aviso.
                    if av.get("sku"):
                        doc["codigo_item_vtex"] = av["sku"]
                    if av.get("produto"):
                        doc["produto"] = av["produto"]
                    if av.get("cod_fornecedor"):
                        doc["codigo_fornecedor"] = av["cod_fornecedor"]
                    if av.get("fornecedor"):
                        doc["departamento"] = av["fornecedor"]
                    # codigo_item_bseller do enrich pode estar de outro item — limpar
                    # se não foi confirmado pelo aviso.
                    if av.get("sku") and not av.get("id_bseller"):
                        # se o enrich trouxe codigo_item_bseller mas o SKU vtex
                        # mudou, esse codigo_item_bseller é de outro item — apagar.
                        if dados_tabelao.get("codigo_item_vtex") and dados_tabelao.get("codigo_item_vtex") != av["sku"]:
                            doc["codigo_item_bseller"] = ""

                    await db.cancelamentos.insert_one(doc)

                    # ANTES do fluxo seguir: busca similar (igual o fluxo manual de
                    # criar_cancelamento faz). Define analise_similar = 'pendente' (tem
                    # similar com estoque) ou 'sem_similar'. Assim o card já abre com
                    # a sugestão de similar pronta pra atendente decidir.
                    try:
                        from routes.produtos_busca import computar_similares_compactos
                        sku_item = doc.get("codigo_item_vtex") or av.get("sku", "")
                        if sku_item:
                            res_sim = await computar_similares_compactos(
                                sku_item, numero_pedido, max_propostos=5
                            )
                            novo_estado = "pendente" if res_sim.get("found") else "sem_similar"
                            await db.cancelamentos.update_one(
                                {"id": novo_id},
                                {"$set": {
                                    "analise_similar": novo_estado,
                                    "similares_sugeridos": res_sim.get("propostos", []),
                                }},
                            )
                            stats["com_similar" if novo_estado == "pendente" else "sem_similar"] = \
                                stats.get("com_similar" if novo_estado == "pendente" else "sem_similar", 0) + 1
                    except Exception as e:
                        logger.warning(
                            f"[smart→aes] busca de similar falhou pro pedido {numero_pedido}: {e}"
                        )

                    await db.avisos_compras.update_one(
                        {"id": av["id"]},
                        {"$set": {
                            "cancelamento_id": novo_id,
                            "atualizado_em": _iso_now_utc(),
                        }},
                    )
                    stats["criados"] += 1
                    logger.info(
                        f"[smart→aes] aviso {av['id']} gerou cancelamento NOVO {novo_id} "
                        f"(pedido {numero_pedido})"
                    )
            except Exception as e:
                stats["erros"] += 1
                logger.exception(f"[smart→aes] falha ao processar aviso {av.get('id')}: {e}")

    except Exception as e:
        logger.exception(f"[smart→aes] falha geral: {e}")
    return stats


# ============== ENDPOINTS ==============

@router.get("/cancelamentos")
async def listar_cancelamentos(
    tipo: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 1000,
    slim: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """Lista cancelamentos com filtros opcionais. Cruza com chamados para indicar atendimentos existentes.

    slim=1: modo leve para o Dashboard ("Demais atividades"). Pula as 5 auto-rotinas
    e os 2 $lookup (chamados/pedidos_erp), retornando só os campos do saldo rolante
    (tipo/status/datas). Reduz a chamada de ~490ms p/ ~80ms e o payload em ~10x.
    """
    if slim:
        proj = {"_id": 0, "tipo": 1, "status": 1, "data_criacao": 1, "data": 1,
                "data_encerramento": 1, "updated_at": 1, "numero_pedido": 1}
        q = {}
        if tipo:
            q["tipo"] = tipo
        if status:
            q["status"] = status
        docs = await db.cancelamentos.find(q, proj).sort("data_criacao", -1).to_list(limit)
        return {"total": len(docs), "cancelamentos": docs}

    # Importa avisos do Smart Compras (anexa em AES existentes ou cria novos)
    try:
        await _auto_importar_avisos_smart_compras()
    except Exception as e:
        logger.warning(f"auto-importar Smart Compras falhou (seguindo com a listagem): {e}")

    # Antes de listar, verifica se há AES pendentes cujo pedido avançou de status (auto-encerramento)
    try:
        await _auto_encerrar_aes_por_status_pedido()
    except Exception as e:
        logger.warning(f"auto-encerrar AES falhou (seguindo com a listagem): {e}")

    # Move automaticamente para a caixa Compras quem tem XD > threshold + pedido ativo + ticket vazio
    try:
        await _auto_mover_para_compras()
    except Exception as e:
        logger.warning(f"auto-mover Compras falhou (seguindo com a listagem): {e}")

    # Nota: para quem tem ticket aberto + XD > threshold, sinaliza pro atendente segurar o cancelamento
    try:
        await _auto_nota_consta_estoque()
    except Exception as e:
        logger.warning(f"auto-nota consta-estoque falhou (seguindo com a listagem): {e}")

    # Auto-volta de Compras: cancelamentos em_compras cujo XD zerou voltam ao canal
    try:
        await _auto_voltar_de_compras()
    except Exception as e:
        logger.warning(f"auto-voltar Compras falhou (seguindo com a listagem): {e}")

    # Reabre AES encerrados que receberam nova solicitação (Smart Compras) após o encerramento
    try:
        await _auto_reabrir_encerrados()
    except Exception as e:
        logger.warning(f"auto-reabrir encerrados falhou (seguindo com a listagem): {e}")

    query = {}
    if tipo:
        query["tipo"] = tipo
    if status:
        query["status"] = status

    # O lookup de estoque (cross-dock) só interessa para AES ainda pendentes — em
    # encerrados o XD não é acionável. Evita varrer estoque_sigeq em listas grandes.
    incluir_estoque = (tipo == "aes") and (status != "encerrado")

    add_fields = {
        "tem_atendimento": {"$gt": [{"$size": "$chamados_relacionados"}, 0]},
        "id_atendimento": {"$arrayElemAt": ["$chamados_relacionados.id_atendimento", 0]},
        "solicitacao_atendimento": {"$arrayElemAt": ["$chamados_relacionados.solicitacao", 0]},
        "pedido_externo": {"$arrayElemAt": ["$pedido_erp_match.pedido_externo", 0]},
        "status_pedido_atual": {"$ifNull": [{"$arrayElemAt": ["$pedido_erp_match.status_pedido", 0]}, ""]},
    }
    project_remove = {"_id": 0, "chamados_relacionados": 0, "pedido_erp_match": 0}

    pipeline = [
        {"$match": query},
        {"$sort": {"data_criacao": -1}},
        {"$limit": limit},
        # Lookup com chamados pra ver se já tem atendimento aberto
        {"$lookup": {
            "from": "chamados",
            "localField": "numero_pedido",
            "foreignField": "numero_pedido",
            "as": "chamados_relacionados",
        }},
        # Lookup com pedidos_erp para trazer pedido_externo (LLL-xxx para LL Loyalty)
        {"$lookup": {
            "from": "pedidos_erp",
            "localField": "numero_pedido",
            "foreignField": "numero_pedido",
            "as": "pedido_erp_match",
        }},
    ]
    if incluir_estoque:
        pipeline.append({"$lookup": {
            "from": "estoque_sigeq",
            "localField": "codigo_item_vtex",
            "foreignField": "cod_terceiro",
            "as": "estoque_xd_match",
        }})
        add_fields["estoque_xd_disp"] = {"$let": {
            "vars": {"m": {"$first": {"$filter": {
                "input": "$estoque_xd_match",
                "cond": {"$eq": ["$$this.source", "SIGEQ425"]},
            }}}},
            "in": {"$ifNull": ["$$m.disp_venda", 0]},
        }}
        project_remove["estoque_xd_match"] = 0

    pipeline.append({"$addFields": add_fields})
    pipeline.append({"$project": project_remove})
    docs = await db.cancelamentos.aggregate(pipeline).to_list(limit)

    # Calcular dias em status para cada um
    for d in docs:
        dt_str = d.get("data_criacao", "")
        try:
            dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
            d["dias_em_status"] = (datetime.now(timezone.utc) - dt).days
        except Exception:
            d["dias_em_status"] = 0
    return {"total": len(docs), "cancelamentos": docs}


def _to_float(v) -> float:
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0


async def _calcular_lifecycle_stats() -> dict:
    """Breakdown por tipo + ciclo de vida (qtd + valor produto+frete), com cache.
    Ciclos (prioridade no doc):
      - não encerrado: tem ticket -> em_tratativa ; sem ticket -> pendente
      - encerrado:     similar (nova entrega c/ SKU novo OU sku_similar) -> similar
                       senão -> cancelado
      - encerrado (card) = similar + cancelado
    Valor = preco_final (produto) + frete da entrega (bigdata)."""
    now = datetime.now(timezone.utc)
    if _lifecycle_cache["ts"] and (now - _lifecycle_cache["ts"]).total_seconds() < _LIFECYCLE_TTL:
        return _lifecycle_cache["data"]

    docs = await db.cancelamentos.find(
        {},
        {"_id": 0, "tipo": 1, "status": 1, "ticket": 1, "sku_similar": 1,
         "preco_final": 1, "numero_pedido": 1, "codigo_item_vtex": 1, "data_criacao": 1},
    ).to_list(None)

    pedidos = list({d.get("numero_pedido") for d in docs if d.get("numero_pedido")})

    # bigdata: frete + pedido_bseller + status por id_entrega; e entregas do master (similar)
    frete_por_entrega = {}
    bseller_por_entrega = {}
    status_por_entrega = {}   # id_entrega -> status BSeller (Cancelado/Liquidado/Aberto)
    entregas_por_master = {}
    try:
        from routes.produtos_busca import _connect_pg
        conn = _connect_pg()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT p.id_entrega, p.pedido_bseller, COALESCE(p.frete, 0), p.status
                    FROM pedidos p WHERE p.id_entrega = ANY(%s)
                    """,
                    (pedidos,),
                )
                masters = set()
                for id_ent, pb, frete, status in cur.fetchall():
                    frete_por_entrega[id_ent] = float(frete or 0)
                    bseller_por_entrega[id_ent] = pb
                    status_por_entrega[id_ent] = status
                    if pb:
                        masters.add(pb)
                if masters:
                    cur.execute(
                        """
                        SELECT p.pedido_bseller, p.id_entrega, p.data_pedido::date, p.status,
                               array_agg(DISTINCT pi.cod_terceiro)
                        FROM pedidos p JOIN pedido_itens pi ON pi.pedido_id = p.id
                        WHERE p.pedido_bseller = ANY(%s)
                        GROUP BY p.pedido_bseller, p.id_entrega, p.data_pedido, p.status
                        """,
                        (list(masters),),
                    )
                    for pb, id_ent, dt, status, skus in cur.fetchall():
                        entregas_por_master.setdefault(pb, []).append({
                            "id_entrega": id_ent, "dt": dt, "status": status,
                            "skus": set(skus or []),
                        })
        finally:
            conn.close()
    except Exception as e:
        logger.warning(f"[lifecycle stats] bigdata indisponível (valor/similar parcial): {e}")

    def tem_similar_bigdata(ped, sku):
        # Similar real só existe se a entrega ORIGINAL foi cancelada (senão o
        # cliente recebeu o original — e SKUs "novos" seriam apenas outros itens
        # do mesmo pedido multi-item, não um substituto).
        if status_por_entrega.get(ped) != "Cancelado":
            return False
        pb = bseller_por_entrega.get(ped)
        entregas = entregas_por_master.get(pb)
        if not pb or not entregas:
            return False
        orig = next((e["skus"] for e in entregas if e["id_entrega"] == ped), {sku})
        dt_orig = next((e["dt"] for e in entregas if e["id_entrega"] == ped), None)
        for e in entregas:
            if e["id_entrega"] == ped or e["status"] == "Cancelado":
                continue
            # Só entrega POSTERIOR ao original conta como reenvio/similar. Entregas
            # de mesma data são itens-irmãos do mesmo pedido multi-item, não substitutos.
            if not (dt_orig and e["dt"] and e["dt"] > dt_orig):
                continue
            if e["skus"] - orig:
                return True
        return False

    def novo_bucket():
        return {
            "pendente": {"n": 0, "valor": 0.0},
            "em_tratativa": {"n": 0, "valor": 0.0},
            "similar": {"n": 0, "valor": 0.0},
            "cancelado": {"n": 0, "valor": 0.0},
            "entregue": {"n": 0, "valor": 0.0},
            "encerrado": {"n": 0, "valor": 0.0},
        }

    out = {"aes": novo_bucket(), "etr": novo_bucket(), "erro_nota": novo_bucket()}
    mensal = {}   # 'YYYY-MM' -> {solicitacoes, similar, cancelado, entregue}
    desde = None
    DESFECHOS = ("similar", "cancelado", "entregue")

    for d in docs:
        tipo = d.get("tipo")
        if tipo not in out:
            continue
        ped = d.get("numero_pedido")
        sku = d.get("codigo_item_vtex")
        valor = _to_float(d.get("preco_final")) + frete_por_entrega.get(ped, 0.0)
        ticket = (d.get("ticket") or "").strip()
        sku_sim = (d.get("sku_similar") or "").strip()

        if d.get("status") != "encerrado":
            # Ainda em aberto: separa por ter ou não ticket
            bucket = "em_tratativa" if ticket else "pendente"
        else:
            # Encerrado → 3 desfechos: Similar (recuperado) / Cancelado (entrega
            # efetivamente cancelada no BSeller) / Entregue (faturado no mesmo
            # código apesar da solicitação — BSeller 'Liquidado').
            # "Recuperado" = similar REALMENTE despachado (bigdata: o master
            # enviou um SKU novo em outra entrega não-cancelada). Ter apenas o
            # sku_similar registrado (proposto) NÃO conta como enviado.
            if tem_similar_bigdata(ped, sku):
                bucket = "similar"
            elif status_por_entrega.get(ped) == "Cancelado":
                bucket = "cancelado"
            else:
                bucket = "entregue"

        out[tipo][bucket]["n"] += 1
        out[tipo][bucket]["valor"] = round(out[tipo][bucket]["valor"] + valor, 2)
        if bucket in DESFECHOS:
            out[tipo]["encerrado"]["n"] += 1
            out[tipo]["encerrado"]["valor"] = round(out[tipo]["encerrado"]["valor"] + valor, 2)

        # --- Série mensal (todos os tipos somados) por mês de data_criacao ---
        dc = d.get("data_criacao") or ""
        mes = str(dc)[:7]
        if len(mes) == 7:
            if desde is None or mes < desde:
                desde = mes
            m = mensal.setdefault(mes, {
                "solicitacoes": 0, "similar": 0, "cancelado": 0, "entregue": 0,
                "v_solicitacoes": 0.0, "v_similar": 0.0, "v_cancelado": 0.0, "v_entregue": 0.0,
            })
            m["solicitacoes"] += 1            # tudo que entrou (volume de solicitações)
            m["v_solicitacoes"] += valor
            if bucket in DESFECHOS:
                m[bucket] += 1
                m["v_" + bucket] += valor

    # Série mensal ordenada
    mensal_lista = []
    for mes in sorted(mensal.keys()):
        mm = mensal[mes]
        mensal_lista.append({
            "mes": mes,
            "solicitacoes": mm["solicitacoes"],
            "similar": mm["similar"],
            "cancelado": mm["cancelado"],
            "entregue": mm["entregue"],
            "v_solicitacoes": round(mm["v_solicitacoes"], 2),
            "v_similar": round(mm["v_similar"], 2),
            "v_cancelado": round(mm["v_cancelado"], 2),
            "v_entregue": round(mm["v_entregue"], 2),
        })

    out["desde"] = desde
    out["mensal"] = mensal_lista

    _lifecycle_cache["ts"] = now
    _lifecycle_cache["data"] = out
    return out


@router.get("/cancelamentos/stats")
async def stats_cancelamentos(current_user: dict = Depends(get_current_user)):
    """Cards de resumo no topo da tela: total/pendentes/encerrados (compat) +
    breakdown por ciclo de vida (pendente/em_tratativa/similar/cancelado) com valor."""
    pipeline_por_tipo = [
        {"$group": {
            "_id": {"tipo": "$tipo", "status": "$status"},
            "count": {"$sum": 1}
        }}
    ]
    resultados = await db.cancelamentos.aggregate(pipeline_por_tipo).to_list(None)
    stats = {
        "aes": {"total": 0, "pendentes": 0, "encerrados": 0},
        "etr": {"total": 0, "pendentes": 0, "encerrados": 0},
        "erro_nota": {"total": 0, "pendentes": 0, "encerrados": 0},
    }
    for r in resultados:
        t = r["_id"].get("tipo")
        s = r["_id"].get("status")
        if t in stats:
            stats[t]["total"] += r["count"]
            if s == "encerrado":
                stats[t]["encerrados"] += r["count"]
            else:
                stats[t]["pendentes"] += r["count"]
    try:
        stats["lifecycle"] = await _calcular_lifecycle_stats()
    except Exception as e:
        logger.warning(f"[stats] lifecycle falhou: {e}")
        stats["lifecycle"] = None
    return stats


@router.get("/cancelamentos/produto-alerta/{sku}")
async def alerta_produto(
    sku: str,
    dias: int = 30,
    current_user: dict = Depends(get_current_user),
):
    """
    Verifica se um produto (por SKU/código terceiro) teve cancelamentos recorrentes.
    Retorna alerta se >3 nos últimos N dias.
    """
    dt_limite = (datetime.now(timezone.utc) - timedelta(days=dias)).isoformat()
    count = await db.cancelamentos.count_documents({
        "$or": [
            {"codigo_item_vtex": sku},
            {"codigo_item_bseller": sku},
        ],
        "data_criacao": {"$gte": dt_limite}
    })
    return {
        "sku": sku,
        "total_cancelamentos": count,
        "periodo_dias": dias,
        "alerta": count > 3,
        "mensagem": f"⚠️ Este produto teve {count} cancelamentos nos últimos {dias} dias. Verificar com Compras/Produção antes de prosseguir." if count > 3 else None
    }


@router.post("/cancelamentos")
async def criar_cancelamento(
    payload: CancelamentoCreate,
    current_user: dict = Depends(get_current_user),
):
    """
    Cria um novo registro de cancelamento.
    Preenche automaticamente dados do pedido a partir do tabelão.
    """
    if payload.tipo not in ("aes", "etr", "erro_nota"):
        raise HTTPException(status_code=400, detail="tipo inválido (use aes, etr ou erro_nota)")

    # Enriquecer com dados do tabelão
    dados_tabelao = await _enrich_from_tabelao(payload.numero_pedido)

    # Verifica alerta de recorrência
    alerta_info = None
    sku = dados_tabelao.get("codigo_item_vtex") or dados_tabelao.get("codigo_item_bseller")
    if sku:
        alerta = await alerta_produto(sku, dias=30, current_user=current_user)
        if alerta.get("alerta"):
            alerta_info = alerta

    user_name = current_user.get("name") or current_user.get("email", "Sistema")
    now = _iso_now_utc()
    now_br = _br_now()

    # Verifica se já existe chamado para esse pedido — se sim, prepend observação
    chamado_existente = await db.chamados.find_one(
        {"numero_pedido": payload.numero_pedido},
        {"_id": 0, "id_atendimento": 1, "solicitacao": 1}
    )
    observacao_final = payload.observacao or ""
    if chamado_existente:
        data_br = now_br.strftime("%d/%m")
        nota = f"{data_br} - já consta no atendimento"
        if nota not in observacao_final:
            observacao_final = (nota + ("\n" + observacao_final if observacao_final else "")).strip()

    doc = {
        "id": str(uuid.uuid4()),
        "tipo": payload.tipo,
        "status": "pendente",
        "numero_pedido": payload.numero_pedido,
        "data_criacao": now,
        "data": now_br.strftime("%Y-%m-%d"),
        "criado_por": user_name,
        "criado_por_email": current_user.get("email", ""),
        # Campos do form
        "motivo": payload.motivo or "",
        "acao": payload.acao or "Cancelar",
        "motivo_rejeicao": payload.motivo_rejeicao or "",
        "ticket": payload.ticket or "",
        "instancia": payload.instancia or "",
        "zerado_reserva": payload.zerado_reserva,
        "observacao": observacao_final,
        # Para erro_nota
        "cliente_confirmado_nome": payload.cliente_confirmado_nome or "",
        "cliente_confirmado_cpf": payload.cliente_confirmado_cpf or "",
        "cliente_confirmado_endereco": payload.cliente_confirmado_endereco or "",
        "nova_entrega": payload.nova_entrega or "",
        # Dados enriquecidos do tabelão
        **dados_tabelao,
        "updated_at": now,
    }

    await db.cancelamentos.insert_one(doc)
    doc.pop("_id", None)

    # AES: busca similares no catálogo (mesma tensão, com estoque) já na criação
    if payload.tipo == "aes":
        try:
            from routes.produtos_busca import computar_similares_compactos
            sku_item = doc.get("codigo_item_vtex") or ""
            res = await computar_similares_compactos(sku_item, doc.get("numero_pedido"), max_propostos=5) if sku_item else {"found": False, "propostos": []}
            novo_estado = "pendente" if res.get("found") else "sem_similar"
            await db.cancelamentos.update_one(
                {"id": doc["id"]},
                {"$set": {"analise_similar": novo_estado, "similares_sugeridos": res.get("propostos", [])}},
            )
            doc["analise_similar"] = novo_estado
            doc["similares_sugeridos"] = res.get("propostos", [])
        except Exception as e:
            logger.warning(f"[criar_cancelamento] busca similar falhou: {e}")

    return {
        "success": True,
        "cancelamento": doc,
        "alerta_produto": alerta_info
    }


@router.put("/cancelamentos/{cancelamento_id}")
async def atualizar_cancelamento(
    cancelamento_id: str,
    payload: CancelamentoUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Atualiza um cancelamento (observação, ticket, instância, status, etc)."""
    update_fields = {k: v for k, v in payload.dict().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="nenhum campo para atualizar")

    update_fields["updated_at"] = _iso_now_utc()

    # Se mudou para encerrado, marca data_encerramento se não veio explícita
    if update_fields.get("status") == "encerrado" and not update_fields.get("data_encerramento"):
        update_fields["data_encerramento"] = _br_now().strftime("%Y-%m-%d")

    # Se sku_similar foi enviado, busca nome e ID do produto no pedidos_erp
    if "sku_similar" in update_fields and update_fields["sku_similar"]:
        sku = update_fields["sku_similar"].strip().upper()
        prod = await db.pedidos_erp.find_one(
            {"codigo_item_vtex": sku},
            {"produto": 1, "codigo_item_bseller": 1, "_id": 0}
        )
        if prod and prod.get("produto"):
            update_fields["nome_similar"] = prod["produto"]
            update_fields["id_similar"] = prod.get("codigo_item_bseller") or ""

    result = await db.cancelamentos.update_one(
        {"id": cancelamento_id},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="cancelamento não encontrado")

    doc = await db.cancelamentos.find_one({"id": cancelamento_id}, {"_id": 0})
    return {"success": True, "cancelamento": doc}


@router.delete("/cancelamentos/{cancelamento_id}")
async def deletar_cancelamento(
    cancelamento_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Remove um cancelamento (uso administrativo)."""
    result = await db.cancelamentos.delete_one({"id": cancelamento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="cancelamento não encontrado")
    return {"success": True}


@router.post("/cancelamentos/enriquecer-atendimentos")
async def enriquecer_atendimentos(current_user: dict = Depends(get_current_user)):
    """
    One-time: para cada cancelamento que tem chamado relacionado em ELO Atendimentos,
    prepend a observação com 'DD/MM - já consta no atendimento' usando a data do cancelamento.
    Não duplica se a nota já existir.
    """
    atualizados = 0
    ja_tinha_nota = 0
    sem_chamado = 0
    cursor = db.cancelamentos.find({}, {"id": 1, "numero_pedido": 1, "observacao": 1, "data": 1, "_id": 0})
    async for c in cursor:
        chamado = await db.chamados.find_one(
            {"numero_pedido": c["numero_pedido"]},
            {"_id": 0, "id_atendimento": 1}
        )
        if not chamado:
            sem_chamado += 1
            continue

        # Data DD/MM da criação do cancelamento (ou hoje se não tiver)
        data_str = c.get("data", "")
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", data_str)
        if m:
            data_br = f"{m.group(3)}/{m.group(2)}"
        else:
            data_br = _br_now().strftime("%d/%m")

        nota = f"{data_br} - já consta no atendimento"
        obs_atual = c.get("observacao", "") or ""
        if "já consta no atendimento" in obs_atual:
            ja_tinha_nota += 1
            continue

        nova_obs = (nota + ("\n" + obs_atual if obs_atual else "")).strip()
        await db.cancelamentos.update_one(
            {"id": c["id"]},
            {"$set": {"observacao": nova_obs}}
        )
        atualizados += 1

    return {
        "atualizados": atualizados,
        "ja_tinha_nota": ja_tinha_nota,
        "sem_chamado": sem_chamado,
    }


@router.get("/cancelamentos/check/{numero_pedido}")
async def check_cancelamentos(
    numero_pedido: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Verifica se um pedido tem cancelamento PENDENTE em AES, ETR ou Erro na Nota.
    Usado no NovoChamado para mostrar alerta ao atendente.
    Cancelamentos já encerrados não disparam o alerta — só os ainda em aberto.
    """
    pedido = str(numero_pedido).strip().split(".")[0]
    # Cancelamento encerrado porque o pedido foi p/ ETR (Entregue a Transportadora):
    # o card mostra um aviso pedindo pra conferir se o cliente foi acionado.
    _mov = await db.cancelamentos.find_one(
        {"numero_pedido": pedido, "status": "encerrado", "encerrado_por_etr": True},
        {"_id": 0, "id": 1})
    aviso_movimentou = bool(_mov)
    # Filtra direto na query: apenas pendentes (ignora encerrados)
    docs = await db.cancelamentos.find(
        {"numero_pedido": pedido, "status": {"$ne": "encerrado"}},
        {"_id": 0, "id": 1, "tipo": 1, "status": 1, "data": 1, "data_encerramento": 1,
         "ticket": 1, "instancia": 1, "observacao": 1, "motivo": 1, "motivo_rejeicao": 1,
         "acao": 1, "criado_por": 1, "nova_entrega": 1}
    ).to_list(20)

    if not docs:
        return {"has_cancelamento": False, "numero_pedido": pedido, "tipos": [],
                "aviso_movimentou": aviso_movimentou}

    # Agrupa por tipo (pega o mais recente de cada — só tem pendentes nesta lista)
    por_tipo = {}
    for d in docs:
        t = d.get("tipo")
        if not t:
            continue
        if t not in por_tipo:
            por_tipo[t] = d

    return {
        "has_cancelamento": True,
        "numero_pedido": pedido,
        "tipos": list(por_tipo.values()),  # lista dos cancelamentos pendentes (1 por tipo)
        "tipos_keys": list(por_tipo.keys()),  # ex: ['aes', 'etr']
        "aviso_movimentou": aviso_movimentou,
    }


@router.post("/cancelamentos/check-lote")
async def check_cancelamentos_lote(payload: dict, current_user: dict = Depends(get_current_user)):
    """Checa EM LOTE quais entregas têm cancelamento PENDENTE (AES/ETR/Erro na Nota) ou
    foram encerradas por movimentação para ETR. Usado pela lista de Atendimentos pra
    mostrar o alerta na linha. payload: {pedidos: [numero_pedido, ...]}."""
    pedidos = list({str(p).strip().split(".")[0] for p in (payload.get("pedidos") or []) if p})
    if not pedidos:
        return {"results": {}}
    results = {}

    def _slot(num):
        return results.setdefault(num, {"has_cancelamento": False, "tipos": [], "aviso_movimentou": False})

    # Pendentes (não encerrados) → alerta "está para cancelamento"
    async for d in db.cancelamentos.find(
        {"numero_pedido": {"$in": pedidos}, "status": {"$ne": "encerrado"}},
        {"_id": 0, "numero_pedido": 1, "tipo": 1},
    ):
        r = _slot(d["numero_pedido"])
        r["has_cancelamento"] = True
        t = d.get("tipo")
        if t and t not in r["tipos"]:
            r["tipos"].append(t)

    # Encerrados por movimentação para ETR → alerta "estava p/ cancelamento, mas movimentou"
    async for d in db.cancelamentos.find(
        {"numero_pedido": {"$in": pedidos}, "status": "encerrado", "encerrado_por_etr": True},
        {"_id": 0, "numero_pedido": 1},
    ):
        _slot(d["numero_pedido"])["aviso_movimentou"] = True

    return {"results": results}


@router.post("/cancelamentos/importar-retorno")
async def importar_retorno_compras(payload: dict, current_user: dict = Depends(get_current_user)):
    """Importa a devolução do Compras (arquivo da Relação Compras).
    payload: {rows: [{entrega, retorno, status}]}. Por linha, casa o AES por numero_pedido:
      - adiciona o 'Retorno Compras' (coluna I) na observação (idempotente);
      - 'Encerrar cancelamento' -> encerra (se já encerrado, só registra a nota);
      - 'Manter cancelamento'   -> mantém pendente + nota; se XD==0 tira da caixa Compras
        (volta pro canal), se XD>0 adiciona nota p/ Compras zerar o estoque e mantém
        (reaparece na próxima Relação Compras, que é AES pendente com XD)."""
    def _n(s):
        return " ".join(str(s or "").lower().split())

    rows = payload.get("rows") or []
    hoje = _br_now().strftime("%d/%m")
    hoje_full = _br_now().strftime("%Y-%m-%d")
    now = _iso_now_utc()
    stats = {"encerrados": 0, "mantidos": 0, "notas": 0, "nao_encontrados": 0, "detalhes": []}

    for row in rows:
        ent = str(row.get("entrega") or "").split(".")[0].strip()
        retorno = str(row.get("retorno") or "").strip()
        status_dec = _n(row.get("status"))
        if not ent:
            continue
        c = await db.cancelamentos.find_one({"numero_pedido": ent, "tipo": "aes"})
        if not c:
            stats["nao_encontrados"] += 1
            stats["detalhes"].append({"entrega": ent, "resultado": "não encontrado (sem AES)"})
            continue

        obs = c.get("observacao", "") or ""
        set_fields = {"updated_at": now}
        if retorno:
            set_fields["retorno_compras"] = retorno
            set_fields["retorno_compras_em"] = hoje_full
            if _n(retorno) not in _n(obs):  # idempotente
                obs = (retorno + ("\n" + obs if obs else "")).strip()
                stats["notas"] += 1

        if "encerrar" in status_dec:
            ja = c.get("status") == "encerrado"
            if not ja:
                set_fields["status"] = "encerrado"
                set_fields["data_encerramento"] = hoje_full
                set_fields["encerrado_por_compras"] = True
            set_fields["em_compras"] = False  # encerrado sai da fila/caixa Compras
            set_fields["observacao"] = obs
            await db.cancelamentos.update_one({"id": c["id"]}, {"$set": set_fields})
            stats["encerrados"] += 1
            stats["detalhes"].append({"entrega": ent, "resultado": "encerrado" + (" (já estava)" if ja else "")})
        else:
            # Manter cancelamento — checa o XD ao vivo (estoque_sigeq SIGEQ425 pelo SKU)
            sku = c.get("codigo_item_vtex") or ""
            xd = 0
            if sku:
                e = await db.estoque_sigeq.find_one({"cod_terceiro": sku, "source": "SIGEQ425"}, {"_id": 0, "disp_venda": 1})
                xd = int((e or {}).get("disp_venda") or 0)
            if xd <= 0:
                set_fields["em_compras"] = False  # sem estoque → volta pro canal cancelar
                res_txt = "mantido · XD=0 → fora da caixa Compras"
            else:
                nota_est = f"{hoje} - Pedido avaliado por Compras, porém consta estoque XD ({xd}) — Compras zerar o estoque"
                if _n(nota_est) not in _n(obs):
                    obs = (nota_est + ("\n" + obs if obs else "")).strip()
                set_fields["em_compras"] = True  # segue na fila → reaparece na próxima Relação Compras
                res_txt = f"mantido · XD={xd} → nota p/ Compras zerar (reaparece na relação)"
            set_fields["observacao"] = obs
            await db.cancelamentos.update_one({"id": c["id"]}, {"$set": set_fields})
            stats["mantidos"] += 1
            stats["detalhes"].append({"entrega": ent, "resultado": res_txt})

    logger.info(f"[importar-retorno] encerrados={stats['encerrados']} mantidos={stats['mantidos']} notas={stats['notas']} nao_enc={stats['nao_encontrados']}")
    return {"success": True, "stats": stats}


@router.get("/cancelamentos/lookup/{numero_pedido}")
async def lookup_pedido(
    numero_pedido: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Busca dados do pedido para preenchimento automático no form de cancelamento.
    Inclui informação de alerta de recorrência por produto.
    """
    dados = await _enrich_from_tabelao(numero_pedido)
    if not dados:
        return {"encontrado": False, "numero_pedido": numero_pedido}

    sku = dados.get("codigo_item_vtex") or dados.get("codigo_item_bseller")
    alerta = None
    if sku:
        a = await alerta_produto(sku, dias=30, current_user=current_user)
        if a.get("alerta"):
            alerta = a

    return {
        "encontrado": True,
        "numero_pedido": numero_pedido,
        "dados": dados,
        "alerta_produto": alerta
    }


@router.get("/cancelamentos/lookup-sku/{sku}")
async def lookup_sku_similar(
    sku: str,
    current_user: dict = Depends(get_current_user),
):
    """Busca produto pelo SKU VTEX: primeiro no MongoDB (pedidos_erp), depois no catálogo PostgreSQL."""
    import os as _os
    import psycopg2
    from psycopg2.extras import RealDictCursor

    sku_clean = sku.strip().upper()

    # 1) MongoDB
    prod = await db.pedidos_erp.find_one(
        {"codigo_item_vtex": sku_clean},
        {"produto": 1, "codigo_item_bseller": 1, "_id": 0}
    )
    if prod and prod.get("produto"):
        return {"found": True, "sku": sku_clean, "nome": prod["produto"],
                "id_bseller": prod.get("codigo_item_bseller") or "", "fonte": "pedidos_erp"}

    # 2) PostgreSQL (itens + item_propriedades)
    pg_host = _os.getenv("PG_HOST")
    if not pg_host:
        raise HTTPException(status_code=404, detail=f"SKU '{sku_clean}' não encontrado no servidor")
    try:
        conn = psycopg2.connect(
            host=pg_host, port=_os.getenv("PG_PORT", "5432"),
            dbname=_os.getenv("PG_DB", "bigdata"), user=_os.getenv("PG_USER"),
            password=_os.getenv("PG_PASSWORD"), connect_timeout=5,
        )
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT DISTINCT ON (ip.cod_terceiro)
                    ip.id_item_bseller::text AS id_bseller,
                    i.descricao AS nome
                FROM item_propriedades ip
                JOIN itens i ON i.cod_terceiro = ip.cod_terceiro
                WHERE UPPER(ip.cod_terceiro) = %s
                LIMIT 1
            """, (sku_clean,))
            row = cur.fetchone()
        conn.close()
    except Exception:
        raise HTTPException(status_code=404, detail=f"SKU '{sku_clean}' não encontrado no servidor")

    if not row:
        raise HTTPException(status_code=404, detail=f"SKU '{sku_clean}' não encontrado no servidor")

    return {"found": True, "sku": sku_clean, "nome": row["nome"],
            "id_bseller": str(row["id_bseller"]) if row["id_bseller"] else "", "fonte": "catalogo"}


@router.post("/cancelamentos/auto-encerrar")
async def auto_encerrar_endpoint(current_user: dict = Depends(get_current_user)):
    """
    Executa manualmente o auto-encerramento de cancelamentos AES cujo pedido
    avançou para 'NFe Aprovada' ou 'Entregue a Transportadora'. Retorna IDs
    encerrados e quantidade.
    """
    return await _auto_encerrar_aes_por_status_pedido()


@router.post("/cancelamentos/analisar-similares")
async def analisar_similares_endpoint(current_user: dict = Depends(get_current_user)):
    """Roda a busca de similares para o passivo de AES pendentes ainda não analisados.
    Manual = força rodada completa (ignora o frescor de 1h usado pelo sync)."""
    return await _analisar_similares_pendentes(limit=500, ignorar_frescor=True)


@router.post("/cancelamentos/{cancelamento_id}/propor-similares")
async def propor_similares(
    cancelamento_id: str,
    payload: dict,
    current_user: dict = Depends(get_current_user),
):
    """
    Analista escolheu um ou mais similares para propor. O item vira 'Similar'
    (acao=Similar), NÃO segue como cancelamento. Os SKUs escolhidos vão para o texto.
    payload: {"skus": ["SKU1", "SKU2", ...]}
    """
    skus = [s.strip().upper() for s in (payload.get("skus") or []) if s and s.strip()]
    if not skus:
        raise HTTPException(status_code=400, detail="Selecione ao menos um similar")

    canc = await db.cancelamentos.find_one({"id": cancelamento_id}, {"_id": 0})
    if not canc:
        raise HTTPException(status_code=404, detail="cancelamento não encontrado")

    sugeridos = {s["sku"]: s for s in (canc.get("similares_sugeridos") or [])}
    escolhidos = [sugeridos[s] for s in skus if s in sugeridos]
    # Se o analista digitou um SKU fora da sugestão, ainda aceitamos (busca rápida no pedidos_erp)
    for s in skus:
        if s not in sugeridos:
            prod = await db.pedidos_erp.find_one({"codigo_item_vtex": s}, {"_id": 0, "produto": 1, "codigo_item_bseller": 1})
            escolhidos.append({
                "sku": s,
                "nome": (prod or {}).get("produto", ""),
                "id_bseller": (prod or {}).get("codigo_item_bseller", "") or "",
                "xd": 0, "ufs": [], "alerta_filial": None,
            })

    # Busca a imagem (VTEX) de cada similar escolhido
    try:
        from routes.produtos_busca import buscar_imagem_vtex
        for e in escolhidos:
            e["image_url"] = await buscar_imagem_vtex(e.get("sku", "")) or ""
    except Exception as ex:
        logger.warning(f"[propor] busca de imagem falhou: {ex}")

    primeiro = escolhidos[0] if escolhidos else {}
    data_br = _br_now().strftime("%d/%m")
    nomes = ", ".join(e["sku"] for e in escolhidos)
    nota = f"{data_br} - Proposto similar(es): {nomes}"

    # Aplica a decisão a TODOS os AES pendentes do mesmo SKU (mesma avaliação serve para todos)
    sku_item = (canc.get("codigo_item_vtex") or "").strip()
    if sku_item:
        irmaos = await db.cancelamentos.find(
            {"tipo": "aes", "status": "pendente", "codigo_item_vtex": sku_item},
            {"_id": 0, "id": 1, "observacao": 1},
        ).to_list(500)
    else:
        irmaos = [{"id": cancelamento_id, "observacao": canc.get("observacao") or ""}]

    for irmao in irmaos:
        obs = irmao.get("observacao") or ""
        nova_obs = obs if nota in obs else (nota + ("\n" + obs if obs else "")).strip()
        await db.cancelamentos.update_one(
            {"id": irmao["id"]},
            {"$set": {
                "acao": "Similar",
                "analise_similar": "proposto",
                "entrou_similar": True,
                "novo": False,
                "similares_propostos": skus,
                # Detalhe (sku, nome, id, imagem) de cada similar escolhido — usado no texto padrão
                "similares_propostos_detalhe": [
                    {"sku": e.get("sku", ""), "nome": e.get("nome", ""),
                     "id_bseller": e.get("id_bseller", ""), "image_url": e.get("image_url", "")}
                    for e in escolhidos
                ],
                "sku_similar": primeiro.get("sku", ""),
                "nome_similar": primeiro.get("nome", ""),
                "id_similar": primeiro.get("id_bseller", ""),
                "observacao": nova_obs,
                "updated_at": _iso_now_utc(),
            }},
        )
    doc = await db.cancelamentos.find_one({"id": cancelamento_id}, {"_id": 0})
    return {"success": True, "cancelamento": doc, "aplicados": len(irmaos)}


@router.post("/cancelamentos/{cancelamento_id}/seguir-cancelamento")
async def seguir_cancelamento(
    cancelamento_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Analista decidiu NÃO propor similar — segue como cancelamento normal (regra atual).
    Aplica a TODOS os AES pendentes do mesmo SKU."""
    canc = await db.cancelamentos.find_one({"id": cancelamento_id}, {"_id": 0})
    if not canc:
        raise HTTPException(status_code=404, detail="cancelamento não encontrado")
    data_br = _br_now().strftime("%d/%m")
    nota = f"{data_br} - Seguir com o cancelamento (sem similar)"

    sku_item = (canc.get("codigo_item_vtex") or "").strip()
    if sku_item:
        irmaos = await db.cancelamentos.find(
            {"tipo": "aes", "status": "pendente", "codigo_item_vtex": sku_item},
            {"_id": 0, "id": 1, "observacao": 1},
        ).to_list(500)
    else:
        irmaos = [{"id": cancelamento_id, "observacao": canc.get("observacao") or ""}]

    for irmao in irmaos:
        obs = irmao.get("observacao") or ""
        nova_obs = obs if "Seguir com o cancelamento" in obs else (nota + ("\n" + obs if obs else "")).strip()
        await db.cancelamentos.update_one(
            {"id": irmao["id"]},
            {"$set": {
                "acao": "Cancelar",
                "analise_similar": "cancelar",
                "novo": False,
                "observacao": nova_obs,
                "updated_at": _iso_now_utc(),
            }},
        )
    doc = await db.cancelamentos.find_one({"id": cancelamento_id}, {"_id": 0})
    return {"success": True, "cancelamento": doc, "aplicados": len(irmaos)}


@router.get("/cancelamentos/relacao-compras")
async def relacao_compras(current_user: dict = Depends(get_current_user)):
    """
    Relação de cancelamentos AES pendentes que TÊM estoque cross-dock
    (XD > threshold) e cujo pedido ainda NÃO foi cancelado.
    Retorna lista com: codigo_fornecedor, produto, sku, xd, ufs_estoque, entrega.
    Usado pelo botão "Relação Compras" da tela de Cancelamentos.
    """
    # 1) AES pendentes com XD > threshold e pedido ainda ativo
    pipeline = [
        {"$match": {"tipo": "aes", "status": "pendente"}},
        {"$lookup": {
            "from": "estoque_sigeq",
            "let": {"sku": "$codigo_item_vtex"},
            "pipeline": [
                {"$match": {"$expr": {"$and": [
                    {"$eq": ["$cod_terceiro", "$$sku"]},
                    {"$eq": ["$source", "SIGEQ425"]},
                ]}}},
                {"$limit": 1},
                {"$project": {"_id": 0, "disp_venda": 1, "codigo_fornecedor": 1, "fornecedor": 1}},
            ],
            "as": "_xd",
        }},
        {"$lookup": {
            "from": "pedidos_erp",
            "localField": "numero_pedido",
            "foreignField": "numero_pedido",
            "as": "_pe",
        }},
        {"$addFields": {
            "xd_disp": {"$ifNull": [{"$arrayElemAt": ["$_xd.disp_venda", 0]}, 0]},
            "cod_fornec": {"$ifNull": [{"$arrayElemAt": ["$_xd.codigo_fornecedor", 0]}, "$codigo_fornecedor"]},
            "fornecedor_nome": {"$ifNull": [{"$arrayElemAt": ["$_xd.fornecedor", 0]}, {"$ifNull": ["$fornecedor", ""]}]},
            "status_pedido_atual": {"$ifNull": [{"$arrayElemAt": ["$_pe.status_pedido", 0]}, ""]},
            "_ticket_vazio": {"$eq": [{"$ifNull": ["$ticket", ""]}, ""]},
        }},
        {"$match": {
            "xd_disp": {"$gt": _XD_THRESHOLD},
            "_ticket_vazio": True,
        }},
        {"$project": {
            "_id": 0,
            "numero_pedido": 1,
            "codigo_item_vtex": 1,
            "produto": 1,
            "cod_fornec": 1,
            "fornecedor_nome": 1,
            "xd_disp": 1,
            "canal_vendas": 1,
            "status_pedido_atual": 1,
            "observacao": 1,
        }},
        {"$sort": {"fornecedor_nome": 1, "produto": 1}},
    ]
    docs = await db.cancelamentos.aggregate(pipeline).to_list(5000)
    if not docs:
        return {"total": 0, "itens": []}

    # 2) Para cada SKU, descobre as UFs onde está o estoque XD (Postgres estoque_xd)
    skus = sorted({d["codigo_item_vtex"] for d in docs if d.get("codigo_item_vtex")})
    ufs_por_sku: dict = {}
    if skus:
        import os as _os
        import psycopg2 as _pg
        from psycopg2.extras import RealDictCursor as _RDC
        try:
            conn = _pg.connect(
                host=_os.getenv("PG_HOST"), port=_os.getenv("PG_PORT", "5432"),
                dbname=_os.getenv("PG_DB", "bigdata"), user=_os.getenv("PG_USER"),
                password=_os.getenv("PG_PASSWORD"), connect_timeout=10,
            )
            with conn.cursor(cursor_factory=_RDC) as cur:
                cur.execute(
                    """
                    SELECT cod_terceiro, filial_id, SUM(disp_venda) AS qtd
                    FROM (
                      SELECT DISTINCT ON (cod_terceiro, filial_id)
                             cod_terceiro, filial_id, disp_venda
                      FROM estoque_xd
                      WHERE UPPER(cod_terceiro) = ANY(%s)
                        AND tipo_deposito = 'XD'
                      ORDER BY cod_terceiro, filial_id, snapshot_date DESC
                    ) t
                    WHERE disp_venda > 0
                    GROUP BY cod_terceiro, filial_id
                    """,
                    ([s.upper() for s in skus],),
                )
                ESTAB_UF = {4: "ES", 5: "ES", 7: "SP", 8: "SP", 10: "SC", 11: "SC"}
                for r in cur.fetchall():
                    sku = r["cod_terceiro"]
                    uf = ESTAB_UF.get(r["filial_id"])
                    if not uf:
                        continue
                    bucket = ufs_por_sku.setdefault(sku, {})
                    bucket[uf] = bucket.get(uf, 0) + int(r["qtd"])
            conn.close()
        except Exception as e:
            logger.warning(f"[relacao-compras] erro no postgres lookup: {e}")

    # 3) Monta saída
    itens = []
    for d in docs:
        sku = d.get("codigo_item_vtex") or ""
        ufs_dict = ufs_por_sku.get(sku.upper()) or {}
        ufs_str = ", ".join(f"{u} ({q})" for u, q in sorted(ufs_dict.items()))
        # "Retorno Compras": se o pedido já foi avaliado por Compras (mas consta estoque),
        # traz a observação já registrada no portal → sinaliza (linha amarela no xlsx).
        obs = d.get("observacao", "") or ""
        linha_aval = next((l.strip() for l in obs.split("\n") if "avaliado por compras" in l.lower()), "")
        itens.append({
            "fornecedor": d.get("fornecedor_nome") or "",
            "codigo_fornecedor": d.get("cod_fornec") or "",
            "produto": d.get("produto") or "",
            "sku": sku,
            "xd": int(d.get("xd_disp") or 0),
            "ufs_estoque": ufs_str or "—",
            "ufs_dict": ufs_dict,  # raw para frontend formatar
            "entrega": d.get("numero_pedido") or "",
            "canal_vendas": d.get("canal_vendas") or "",
            "status_pedido": d.get("status_pedido_atual") or "",
            "retorno_compras": linha_aval,
            "ja_avaliado": bool(linha_aval),
        })
    return {"total": len(itens), "itens": itens}


@router.get("/cancelamentos/relacao-compras-xlsx")
async def relacao_compras_xlsx(current_user: dict = Depends(get_current_user)):
    """Gera a Relação Compras já em .xlsx (com a coluna 'Retorno Compras' e as linhas
    já avaliadas pelo Compras destacadas em amarelo). Feito no backend porque a
    geração no front (SheetJS free) não pinta células."""
    from fastapi.responses import StreamingResponse
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    data = await relacao_compras(current_user)
    itens = data.get("itens", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Relação de Compras - AES"
    cols = ["Fornecedor", "Cód. fornecedor", "Produto", "SKU", "Estoque XD",
            "UF do estoque", "Entrega", "Parceiro", "Retorno Compras"]
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    amarelo = PatternFill("solid", fgColor="FFFF00")
    for it in itens:
        ws.append([
            it.get("fornecedor", ""), it.get("codigo_fornecedor", ""), it.get("produto", ""),
            it.get("sku", ""), it.get("xd", 0), it.get("ufs_estoque", ""),
            it.get("entrega", ""), it.get("canal_vendas", ""), it.get("retorno_compras", ""),
        ])
        if it.get("ja_avaliado"):
            for cell in ws[ws.max_row]:
                cell.fill = amarelo
    for i, w in enumerate([22, 16, 55, 12, 12, 14, 12, 14, 55], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Relacao de Compras - AES {_br_now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/cancelamentos/similar-custo")
async def similar_custo_diff(current_user: dict = Depends(get_current_user)):
    """Similares REALMENTE enviados (recuperados) — original × similar despachado.
    Mesma detecção do card "recuperado": via bigdata, o master (pedido_bseller)
    despachou um SKU novo numa entrega não-cancelada (≠ a original, mesma data ou
    depois). O SKU exibido é o que FOI despachado — se diferente do sku_similar
    registrado no chamado, prevalece o real. dif_media = média de (custo do similar
    − custo do original); positivo = similar saiu mais caro; negativo = economia."""
    from routes.produtos_busca import _connect_pg
    docs = await db.cancelamentos.find(
        {"tipo": "aes", "status": "encerrado"},
        {"_id": 0, "numero_pedido": 1, "produto": 1, "codigo_item_vtex": 1,
         "sku_similar": 1, "nome_similar": 1, "canal_vendas": 1, "data": 1, "preco_final": 1},
    ).to_list(20000)
    entregas = [str(d.get("numero_pedido") or "").split(".")[0] for d in docs if d.get("numero_pedido")]

    bseller, ent_master, custo, status_orig = {}, {}, {}, {}
    raw = []
    itens = []
    difs, cos, css = [], [], []
    if entregas:
        try:
            conn = _connect_pg()
            try:
                with conn.cursor() as cur:
                    # master (pedido_bseller) de cada entrega
                    cur.execute("SELECT id_entrega, pedido_bseller, status FROM pedidos WHERE id_entrega = ANY(%s)", (entregas,))
                    masters = set()
                    for ie, pb, st in cur.fetchall():
                        bseller[ie] = pb
                        status_orig[ie] = st
                        if pb:
                            masters.add(pb)
                    # entregas de cada master (data + status + SKUs) p/ detectar o despacho
                    if masters:
                        cur.execute(
                            """
                            SELECT p.pedido_bseller, p.id_entrega, p.data_pedido::date, p.status,
                                   array_agg(DISTINCT pi.cod_terceiro)
                            FROM pedidos p JOIN pedido_itens pi ON pi.pedido_id = p.id
                            WHERE p.pedido_bseller = ANY(%s)
                            GROUP BY p.pedido_bseller, p.id_entrega, p.data_pedido, p.status
                            """,
                            (list(masters),),
                        )
                        for pb, ie, dt, st, sks in cur.fetchall():
                            ent_master.setdefault(pb, []).append({
                                "id_entrega": ie, "dt": dt, "status": st, "skus": set(sks or []),
                            })

                    def _similar_enviado(ent, orig_sku, rec_sim):
                        # SKU do similar REALMENTE despachado (bigdata). Regras (idênticas
                        # ao card "recuperado"): a original tem que estar CANCELADA e o
                        # substituto vem numa entrega POSTERIOR não-cancelada com SKU novo.
                        # Entregas de mesma data são itens-irmãos (multi-item), não similar.
                        if status_orig.get(ent) != "Cancelado":
                            return None
                        pb = bseller.get(ent)
                        ents = ent_master.get(pb)
                        if not pb or not ents:
                            return None
                        orig = next((e["skus"] for e in ents if e["id_entrega"] == ent), {orig_sku})
                        dt_orig = next((e["dt"] for e in ents if e["id_entrega"] == ent), None)
                        candidatos = []
                        for e in ents:
                            if e["id_entrega"] == ent or e["status"] == "Cancelado":
                                continue
                            if not (dt_orig and e["dt"] and e["dt"] > dt_orig):
                                continue
                            candidatos.extend(e["skus"] - orig)
                        if not candidatos:
                            return None
                        recU = (rec_sim or "").upper()
                        for c in candidatos:  # se algum bate com o registrado, prioriza
                            if c.upper() == recU:
                                return c
                        return sorted(candidatos)[0]

                    skus_custo = set()
                    for d in docs:
                        ent = str(d.get("numero_pedido") or "").split(".")[0]
                        orig = (d.get("codigo_item_vtex") or "").strip()
                        if not (ent and orig):
                            continue
                        ship = _similar_enviado(ent, orig, (d.get("sku_similar") or "").strip())
                        if not ship:
                            continue
                        raw.append((d, ent, orig, ship))
                        skus_custo.add(orig.upper())
                        skus_custo.add(ship.upper())

                    if skus_custo:
                        cur.execute(
                            """
                            SELECT UPPER(cod_terceiro), preco FROM (
                                SELECT DISTINCT ON (cod_terceiro) cod_terceiro, preco
                                FROM precos_compra_hist
                                WHERE UPPER(cod_terceiro) = ANY(%s) AND preco > 0
                                ORDER BY cod_terceiro, data_alteracao DESC NULLS LAST
                            ) t
                            """,
                            (list(skus_custo),),
                        )
                        for cod, preco in cur.fetchall():
                            custo[cod] = float(preco or 0)
            finally:
                conn.close()
        except Exception as e:
            logger.warning(f"[similar-custo] erro no postgres: {e}")
            raw = []

    for d, ent, orig, ship in raw:
        rec = (d.get("sku_similar") or "").strip()
        divergente = bool(rec) and rec.upper() != ship.upper()
        co = custo.get(orig.upper())
        cs = custo.get(ship.upper())
        dif = round(cs - co, 2) if (co and cs and co > 0 and cs > 0) else None
        if dif is not None:
            difs.append(dif)
            cos.append(co)
            css.append(cs)
        itens.append({
            "entrega": ent,
            "produto": d.get("produto") or "",
            "sku_original": orig,
            "sku_similar": ship,                       # o que FOI despachado
            "sku_registrado": rec if divergente else "",  # registrado no chamado, se diferente
            "nome_similar": (d.get("nome_similar") or "") if not divergente else "",
            "canal": d.get("canal_vendas") or "",
            "data": d.get("data") or "",
            "custo_original": round(co, 2) if (co and co > 0) else None,
            "custo_similar": round(cs, 2) if (cs and cs > 0) else None,
            "diferenca": dif,
            "preco_venda": round(_to_float(d.get("preco_final")), 2) if d.get("preco_final") else None,
        })
    # maiores diferenças primeiro; itens sem custo conhecido vão pro fim
    itens.sort(key=lambda x: (x["diferenca"] is None, -(x["diferenca"] if x["diferenca"] is not None else 0)))
    n = len(difs)
    return {
        "n": n,
        "total_pares": len(itens),
        "dif_media": round(sum(difs) / n, 2) if n else 0.0,
        "custo_orig_medio": round(sum(cos) / n, 2) if n else 0.0,
        "custo_sim_medio": round(sum(css) / n, 2) if n else 0.0,
        "itens": itens,
    }


@router.get("/cancelamentos/custo-skus")
async def custo_skus(skus: str = "", current_user: dict = Depends(get_current_user)):
    """Último preço de compra (custo) de cada SKU. Query: skus=A,B,C → {custos: {SKU: valor}}."""
    from routes.produtos_busca import _connect_pg
    lista = [s.strip() for s in (skus or "").split(",") if s.strip()]
    if not lista:
        return {"custos": {}}
    custos = {}
    try:
        conn = _connect_pg()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT UPPER(cod_terceiro), preco FROM (
                        SELECT DISTINCT ON (cod_terceiro) cod_terceiro, preco
                        FROM precos_compra_hist
                        WHERE UPPER(cod_terceiro) = ANY(%s) AND preco > 0
                        ORDER BY cod_terceiro, data_alteracao DESC NULLS LAST
                    ) t
                    """,
                    ([s.upper() for s in lista],),
                )
                for cod, preco in cur.fetchall():
                    custos[cod] = round(float(preco or 0), 2)
        finally:
            conn.close()
    except Exception as e:
        logger.warning(f"[custo-skus] erro no postgres: {e}")
    return {"custos": custos}
