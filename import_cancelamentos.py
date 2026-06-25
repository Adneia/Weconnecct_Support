import asyncio, os, sys, uuid
import openpyxl
from datetime import datetime, timezone

sys.path.insert(0, '/app')
os.environ['MONGO_URL'] = 'mongodb://elo-mongo-test:27017'
os.environ['DB_NAME'] = 'elo_weconnect'

def fmt_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    return s if s else None

def fmt_str(val):
    if val is None:
        return ''
    return str(val).strip()

def fmt_entrega(val):
    if val is None:
        return None
    try:
        return str(int(float(str(val).strip())))
    except Exception:
        s = str(val).strip()
        return s if s else None

async def main():
    from utils.database import db

    wb = openpyxl.load_workbook('/tmp/cancelamento_geral.xlsx', read_only=True, data_only=True)
    now = datetime.now(timezone.utc).isoformat()

    totals = {}

    # ============================================================
    # AES
    # Cols: Data(0) Entrega(1) CodTerceiro(2) Parceiro(3) Ação(4)
    #       Ticket(5) Instancia(6) ZeradoReserva(7) StatusBseller(8)
    #       DataEnc(9) Observacao(10)
    # ============================================================
    ws = wb['AES']
    upd = ins = skip = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        entrega = fmt_entrega(row[1])
        if not entrega:
            skip += 1
            continue

        data_enc = fmt_date(row[9])
        status = 'encerrado' if data_enc else 'aberto'

        upd_fields = {
            'acao':                   fmt_str(row[4]),
            'ticket':                 fmt_str(row[5]),
            'instancia':              fmt_str(row[6]),
            'zerado_reserva_planilha': fmt_str(row[7]),
            'status_bseller_planilha': fmt_str(row[8]),
            'data_encerramento':      data_enc,
            'observacao':             fmt_str(row[10]),
            'status':                 status,
            'updated_at':             now,
        }

        res = await db.cancelamentos.update_one(
            {'tipo': 'aes', 'numero_pedido': entrega},
            {'$set': upd_fields}
        )

        if res.matched_count > 0:
            upd += 1
        else:
            doc = {
                'tipo':                    'aes',
                'numero_pedido':           entrega,
                'data':                    fmt_date(row[0]),
                'data_criacao':            fmt_date(row[0]),
                'codigo_terceiro_planilha': fmt_str(row[2]),
                'parceiro_planilha':       fmt_str(row[3]),
                **upd_fields,
                'id':           str(uuid.uuid4()),
                'criado_por':   'Importação Planilha',
                'criado_por_email': '',
            }
            await db.cancelamentos.insert_one(doc)
            ins += 1

    totals['AES'] = {'atualizado': upd, 'inserido': ins, 'pulado': skip}
    print(f'AES: atualizado={upd}, inserido={ins}, pulado(sem entrega)={skip}')

    # ============================================================
    # ETR
    # Cols: Data(0) Entrega(1) Parceiro(2) Nota(3) Motivo(4) Ação(5)
    #       Ticket(6) Instancia(7) DataEnc(8) Observacao(9)
    # ============================================================
    ws = wb['ETR']
    upd = ins = skip = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        entrega = fmt_entrega(row[1])
        if not entrega:
            skip += 1
            continue

        data_enc = fmt_date(row[8])
        status = 'encerrado' if data_enc else 'aberto'

        upd_fields = {
            'parceiro_planilha':    fmt_str(row[2]),
            'nota_fiscal_planilha': fmt_str(row[3]),
            'motivo':               fmt_str(row[4]),
            'acao':                 fmt_str(row[5]),
            'ticket':               fmt_str(row[6]),
            'instancia':            fmt_str(row[7]),
            'data_encerramento':    data_enc,
            'observacao':           fmt_str(row[9]),
            'status':               status,
            'updated_at':           now,
        }

        res = await db.cancelamentos.update_one(
            {'tipo': 'etr', 'numero_pedido': entrega},
            {'$set': upd_fields}
        )

        if res.matched_count > 0:
            upd += 1
        else:
            doc = {
                'tipo':         'etr',
                'numero_pedido': entrega,
                'data':         fmt_date(row[0]),
                'data_criacao': fmt_date(row[0]),
                **upd_fields,
                'id':           str(uuid.uuid4()),
                'criado_por':   'Importação Planilha',
                'criado_por_email': '',
            }
            await db.cancelamentos.insert_one(doc)
            ins += 1

    totals['ETR'] = {'atualizado': upd, 'inserido': ins, 'pulado': skip}
    print(f'ETR: atualizado={upd}, inserido={ins}, pulado(sem entrega)={skip}')

    # ============================================================
    # ERRO NA NOTA
    # Cols: Data(0) Nota(1) Filial(2) Entrega(3) MotivoRejeição(4)
    #       Status/Obs(5) NovaEntrega(6) Instancia(7) Encerrado(8)
    # ============================================================
    ws = wb['Erro na Nota']
    upd = ins = skip = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        entrega = fmt_entrega(row[3])
        if not entrega:
            skip += 1
            continue

        data_enc = fmt_date(row[8])
        status = 'encerrado' if data_enc else 'aberto'

        upd_fields = {
            'nota_fiscal_planilha': fmt_str(row[1]),
            'filial_planilha':      fmt_str(row[2]),
            'motivo_rejeicao':      fmt_str(row[4]),
            'observacao':           fmt_str(row[5]),   # "Status" col = notas de andamento
            'nova_entrega':         fmt_entrega(row[6]) or '',
            'instancia':            fmt_str(row[7]),
            'data_encerramento':    data_enc,
            'status':               status,
            'updated_at':           now,
        }

        res = await db.cancelamentos.update_one(
            {'tipo': 'erro_nota', 'numero_pedido': entrega},
            {'$set': upd_fields}
        )

        if res.matched_count > 0:
            upd += 1
        else:
            doc = {
                'tipo':         'erro_nota',
                'numero_pedido': entrega,
                'data':         fmt_date(row[0]),
                'data_criacao': fmt_date(row[0]),
                **upd_fields,
                'id':           str(uuid.uuid4()),
                'criado_por':   'Importação Planilha',
                'criado_por_email': '',
            }
            await db.cancelamentos.insert_one(doc)
            ins += 1

    totals['Erro Nota'] = {'atualizado': upd, 'inserido': ins, 'pulado': skip}
    print(f'Erro Nota: atualizado={upd}, inserido={ins}, pulado(sem entrega)={skip}')

    wb.close()

    total_upd = sum(v['atualizado'] for v in totals.values())
    total_ins = sum(v['inserido'] for v in totals.values())
    print(f'\n✅ CONCLUÍDO — Total: {total_upd} atualizados, {total_ins} inseridos')

asyncio.run(main())
